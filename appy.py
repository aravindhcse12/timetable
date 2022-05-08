
#from calendar import c
#from shutil import register_archive_format
#from ssl import ALERT_DESCRIPTION_BAD_CERTIFICATE_HASH_VALUE
#from unicodedata import name
#from webbrowser import get
#from flask import Flask, render_template, request, redirect, url_for, flash,jsonify,session
#pip install flask_mail

from flask import *  

from flask_sqlalchemy import SQLAlchemy,BaseQuery
from werkzeug.security import generate_password_hash, check_password_hash

import json
from importlib_metadata import pass_none

#from mysqlx import Session
from sqlalchemy import null,desc

from openpyxl import Workbook

from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import datetime

import time

import random
from flask_mail import Mail,Message
#from insert.py file
from insert import updateF,deleteF,show_all_tables,create_table,mycursor,mydb

app = Flask(__name__)
app.secret_key = "Secret Key"
 
#SqlAlchemy Database Configuration With Mysql
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql://root:''@localhost/jntuk1'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
 
db = SQLAlchemy(app)

total_list=[]
#mysql.connector connection code

"""mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  password="",
  database="jntuk"
)
"""

#mail variables
with open('config.json','r') as f:
    credentials=json.load(f)['credentials']

app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] =465
app.config['MAIL_USERNAME'] = credentials['user-id']
app.config['MAIL_PASSWORD'] =credentials['pwd']
app.config['MAIL_DEFAULT_SENDER'] = credentials['user-id']
app.config['MAIL_USE_TLS'] = False
app.config['MAIL_USE_SSL'] = True
mail=Mail(app)

def sendmail(subject,recipients,body):
    print("HAII")
    print(credentials)
    try:

        """msg = Message(
                'Hello',
                sender ='aravindhchalla@gmail.com',
                recipients = ['dsubbu7661@gmail.com']
               )
        msg.body = 'Hello Flask message sent from Flask-Mail'

        mail.send(msg)"""
        msg=Message()   
        msg.subject=subject
        msg.recipients=[recipients]
        msg.html =body

        print(msg.html,msg.recipients,msg.subject)
        mail.send(msg)
        

        print(mail.send(msg))

    except Exception as e:
        print("Exception-",str(e))
        return

@app.route('/exclude',methods = ['GET', 'POST'])
def exclude():
    get_data=request.get_json()

    print(get_data)

    fac=Teacheradd.query.get(int(get_data['ids']))

    if(fac.exclude==0):
        fac.exclude=1
        db.session.commit()
    else:
        fac.exclude=0
        db.session.commit()

    return jsonify({'data':'success'})

@app.route('/forgotpassword',methods = ['GET', 'POST'])
def forgotpassword():
    if not 'loggedin' in session:

        print("reset successful")
        get_data=request.get_json()
        if request.method == 'POST' and 'email' in get_data:
            try:
                otp=random.randint(1000,9999)
                print(otp,get_data['email'])

                #print(request.form['email'])

                mycursor.execute("SELECT * FROM login WHERE username='{0}'".format(get_data['email']))

                account = mycursor.fetchone()

                print(account)

                if(account):

                    print("before update")
                    mycursor.execute("UPDATE login SET otp={0}  WHERE username='{1}'".format(otp,get_data['email']))
                    mydb.commit() 

                    print("after update")
                    print(get_data['email'])
                    sendmail(" here is your OTP",get_data['email'],str(otp))   
                    
                    return jsonify({'otp' :'success'})
                else:
                    return jsonify({'msg':'Email does not Exist'})

            except Exception as e:
                print(str(e))
                return jsonify({'msg':'SOMETHING WRONG'})
        elif(request.method=='POST'  and 'otp' in get_data):
            try:
                
                mycursor.execute("SELECT otp FROM login where otp={0}".format(get_data['otp']))
                otpAccount = mycursor.fetchone()

                print(otpAccount)
                if int(int(get_data['otp']))==int(otpAccount[0]):

                    if get_data['cnf']!=get_data['pass']:

                        return  jsonify({'msg' :"Password and Confirm Password Doesn't Match"})

                    print("Password and Confirm Password  Match")

                    mycursor.execute("UPDATE login SET password='{0}' WHERE otp={1}".format(generate_password_hash(get_data['pass']),get_data['otp']))

                    mydb.commit()
                
                    mycursor.execute("UPDATE login SET otp='{0}' WHERE otp={1}".format(0,get_data['otp']))

                    mydb.commit()

                    
                    flash("Password Updated Successfully",'success')

                    return redirect(url_for('dashboard'))

                return jsonify({'error' :1})
            except Exception as e: 
                return  jsonify({'msg' :'SOMETHING WRONG AT OTP'})
        else:
            return jsonify({'msg':'Invalid Data'}) 


@app.route('/changepassword',methods = ['GET', 'POST'])
def changepassword():
    if 'loggedin' in session:

        return render_template("changepassword.html")
    else:
        return render_template("login.html")

@app.route('/change_faculty_password',methods = ['GET', 'POST'])
def change_faculty_password():
    if 'loggedin' in session:
        return render_template("change_faculty_password.html")
    else:
        return render_template("login.html")

@app.route('/change_password_funtion/<choice>',methods = ['GET', 'POST'])
def change_password_funtion(choice):
    
    if 'loggedin' in session:
        

        if request.method == 'POST':
            
            username = request.form['email']

            old = request.form['old']

            new=request.form['new']

            check=login.query.filter_by(username=username).first()

            if(check):

                if(session["mail"]!=username):

                    flash("Mail is invalid")

                    if(choice!='faculty'):
                        return render_template("changepassword.html")
                    else:
                        return render_template("change_faculty_password.html")

                print(username,old)

                print("decrypt",check_password_hash(check.password,old))

                if(check_password_hash(check.password,old)):
                    #update password

                    if(new==old):

                        flash("New password  is same as old password")

                        if(choice!='faculty'):
                            return render_template("changepassword.html")
                        else:
                            return render_template("change_faculty_password.html")

                    check.password=generate_password_hash(new)

                    db.session.commit()

                    flash("Password updated successfully",'success')

                    return redirect(url_for('logout'))
                    
                else:
                    flash("invalid password",'danger')
                    if(choice!='faculty'):
                       
                        return render_template('changepassword.html')
                    else:
                        
                        return render_template("change_faculty_password.html")

            else:
                flash("Mail Does not exist",'danger')
                if(choice!="faculty"):
                    
                    return render_template("changepassword.html")
                else:
                    return render_template("change_faculty_password.html")
    else:

        return render_template("login.html")

        
@app.route('/reset_password')
def reset_password():
    if 'loggedin' not in session:

        #flash("reset successful")

        return render_template("forgotPassword.html")
    else:

        return redirect("/")
#This is the dashboard route where we are going to
@app.route('/')
def dashboard():
    #session.clear()

    if 'loggedin' in session:

    
        print("login session already there")

        print(*session)
        print("dashboard")

        get_day_period=day_period.query.get(1)
        if(get_day_period):

            print("something")

        else:
            print("nothing")

            default_day_period=day_period(6,7,14,14,16,16,4)
            db.session.add(default_day_period)
            db.session.commit()    

        #return render_template("login.html")
        if(session['type']=='faculty_type'):
            
            return redirect(url_for('faculty_table'))
        else:
            return redirect(url_for('classwisetimetable'))


    else:#not logged in 
        return render_template("login.html")


@app.route('/login_check', methods = ['GET', 'POST'])
def login_check():
    session.clear()

    if 'loggedin' not in session:


        if request.method == 'POST':
    
            username = request.form['email']

            password = request.form['password']

            check=login.query.filter_by(username=username).first()

            print(username,password)

            print(check)

            if(check):

                print(username,password)

                print("decrypt",check_password_hash(check.password, password))

                """  print("principal password",generate_password_hash("principal_type"))

                mail="principaljntuk@gmail.com"

                pasw='principal@jntuk'

                login_data=login(mail,generate_password_hash(pasw),"dept_type")
                db.session.add(login_data)
                db.session.commit()"""


                if(check_password_hash(check.password, password)):

                    session['loggedin']=1

                    session['mail']=username


                    print("check.type",check.type)
                    if(check.type=='faculty_type'):

                        session['type']='faculty_type'
                        return redirect(url_for('faculty_table'))              

                    elif(check.type=='dept_type'):
                        session['type']='dept_type'
                        return redirect(url_for('classwisetimetable'))
                    else:
                        session['type']='principal_type'
                        return redirect(url_for('classwisetimetable'))
                else:
                    flash("invalid password")
                    return redirect(url_for('dashboard'))
                            

            else:

                flash("mail does not exists")

                return redirect(url_for('dashboard'))

    else:
        return redirect(url_for('dashboard'))


@app.route('/faculty_table', methods = ['GET', 'POST'])
def faculty_table():

    if('loggedin' in session):

          #collect faculty data

        data=Teacheradd.query.filter_by(mail=session['mail']).first()

        #collect faculty time table

        mycursor.execute("SELECT * FROM `{0}`".format(data.name))
        fac=mycursor.fetchall()


        des=classconfig_table.query.all()

        periods=day_period.query.get(1)


        return render_template("faculty_table.html",faculty=data.name,dept=data.branch,work_load=data.work_load,designation=data.role,des=des,fac=fac,periods=periods.periods)
    else:
        return redirect(url_for('dashboard'))

@app.route('/logout', methods = ['GET', 'POST'])
def logout():   

    

    if('loggedin' in session):

        session.clear()

    return redirect(url_for('dashboard'))


#Creating Teacheradd table for our jntuk database
class login(db.Model):

    __tablename__ = "login"

    id = db.Column(db.Integer, primary_key = True)
    username = db.Column(db.String(100))
    password = db.Column(db.String(255))
    type=db.Column(db.String(50))
    otp=db.Column(db.Integer)

    def __init__(self, username,password,type,otp):
        self.username = username
        self.password=password
        self.type=type
        self.otp=otp

        
db.create_all()
@app.route('/generate_class', methods = ['GET', 'POST'])
def generate_class():

    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))        

        mycursor = mydb.cursor(buffered=True)


        get_data=request.get_json()

        msg=""

        #For lab configuration of single class

        #get_data=request.get_json()

        get_class_ele=classconfig_table.query.filter_by(classname=get_data['data'],type="LAB").all()


        get_session=day_period.query.get(1)

        already_lab_day=[] #on successful updation of single day add day to this of split

        for row in  get_class_ele:
            print(row.subject)
            print(row.faculty)

            if(row.total_periods == row.allocated):
                continue

            if(row.faculty!="NA"):

                msg="{0} subject is already been assigned with faculty eiether you remove the faculty or do manually".format(row.subject)
                continue
                
            split_slash=row.split.split('/')

            get_split=[]

            print(split_slash)
            
            if(len(split_slash)>1):
                
                for pc in split_slash:
                        get_split.append(pc.split(','))

                for j in get_split[1]:
                        if(j in get_split[0]):
                                    get_split[0].remove(j)

                get_split=get_split[0]

                print(get_split)
            else:

                get_split=split_slash[0].split(',')

            fac_lab_list=[]

            update_ready=[]

            flag_overflow=0

            split_count=0

            collect_split=[]
            print(row.subject,"split",get_split)
            for sp in get_split:

                #find the best fit of split eiether morning or afternoon

                morning=int(get_session.morning_periods)-int(sp)

                print("morning",morning)

                afternoon=int(get_session.periods-get_session.morning_periods)-int(sp)
                print("afternoon",afternoon)

                flag_both=0

                if(morning < afternoon and morning>-1):
                    #select morning
                    print("morning if")
                    flag=0
                elif(afternoon < morning and afternoon>-1):
                    #select afternoon
                    print("afternoon elif")
                    flag=1

                else:
                    print("else random")
                    if(afternoon>-1 and morning>-1):
                        print("afternoon elif")
                        flag=random.choice([0,1])
                        flag_both=1
                    elif(morning>-1):
                        flag=0
                    else:
                        flag=1


                print("flag",flag)

                #generate strings of periods
                get_day=day_period.query.get(1)

                if(flag==0):

                    start=1
                    print("morning selected start:",start)
                    
                else:#afternoon
                    start=int(get_day.morning_periods)+1
                    print("afternoon selected start:",start)

                #collecting days

                get_day=day_period.query.get(1)
                day=int(get_day.day)

                print("max day:",day)

                day_list=['MON','TUE','WED','THU','FRI','SAT','SUN']   

        
                day_list=day_list[:day]

                random.shuffle(day_list)

                period_colloction_final=[]

                if(flag_both==1):

                    for start in list([1,get_day.morning_periods+1]):
                         for d in day_list:

                                        
                            period_collection=[]

                            if(d in already_lab_day):
                                continue

                            for start_scale in range(start,start+int(sp)):

                                period_collection.append(d+"--"+str(start_scale))

                            period_colloction_final.append(period_collection)
                else:

                    for d in day_list:

                                        
                        period_collection=[]

                        if(d in already_lab_day):
                            continue


                        for start_scale in range(start,start+int(sp)):

                            period_collection.append(d+"--"+str(start_scale))

                        period_colloction_final.append(period_collection)



                if(period_colloction_final):
                    print("period collection_final:",period_colloction_final)
                else:

                    flag_overflow=1
                    print("do it manually no slot found")
                    break

                #get head faculty

                
                index=row.subject.find(" LAB")

                sub_theory=row.subject[:index]

                mycursor.execute("SELECT faculty FROM classconfig_table WHERE subject='{0}' AND type='{2}' AND classname='{1}'".format(sub_theory,row.classname,"THEORY"))

                val=mycursor.fetchall()

                if(val):
                    print("subject theory for lab head faculty:",val[0][0])
                    get_head=Teacheradd.query.filter_by(name=val[0][0]).first()

                    
                    mycursor.execute("SELECT total_Periods,allocated FROM classconfig_table WHERE faculty='{0}' AND type='{2}'".format(val[0][0],row.classname,"THEORY"))

                    summ_check=mycursor.fetchall()

                    summ=0
                    for p in summ_check:

                        print(p)
                        #this summm must to be scheduled
                        summ+=int(p[0])-int(p[1])

                    designation_dic={'Professor':'proffesor','Associate Professor':'Assoc_prof','Assistant Professor':'Asst_prof','Assistant Professor(C)':'Asst_prof_c'}                        


                    mycursor.execute("select `{0}` FROM day_periods WHERE id={1}".format(designation_dic[get_head.role],1))

                    designation=mycursor.fetchall()

                    #day_period.query.get(1)

                    
                    if(get_head.work_load+int(sp)+summ<=int(designation[0][0]) ):


                        print(val[0][0],get_head.work_load+int(sp)+summ," is less than his/her workload can procced")

                        if(val[0][0] not  in fac_lab_list):
                            fac_lab_list.append(val[0][0])
                    else:

                        if(get_head.role=="Assistant Professor(C)"):
                            fac_lab_list.append(val[0][0])
                        else:   
                            print(val[0][0],"head faculty is buzy")
                    
                        
                else:
                    
                    print("No subject found,so no head faculty")

                for p in period_colloction_final:

                    flag_select=0

                    for row_p in p:
                    
                        period_split=row_p.split("--")
                        #select class
                        mycursor.execute("SELECT `{1}` FROM `{0}` WHERE DAY='{2}';".format(get_data['data'],period_split[1],period_split[0]))

                        val1=mycursor.fetchall()

                        if(val1[0][0]!="--"):
                            flag_select=1

                            break   

                        #select lab

                        all_lab=row.lab.split(',')
                        for lab in all_lab:

                            mycursor.execute("SELECT `{1}` FROM `{0}` WHERE DAY='{2}';".format(lab,period_split[1],period_split[0]))

                            val1=mycursor.fetchall()
                    
                            if(val1[0][0]!="--"):

                                flag_select=1

                                break   


                        #select fac
                        print("faculty checking slot",row_p)
                        if(len(fac_lab_list)>1):
                            mycursor.execute("SELECT `{1}` FROM `{0}` WHERE DAY='{2}';".format(fac_lab_list[0],period_split[1],period_split[0]))

                            val1=mycursor.fetchall()
                    
                            if(val1[0][0]!="--"):
                                flag_select=1

                                break 

                    if(flag_select):
                        #updated_day=p[0].split('--')
                        #already_lab_day.append(updated_day[0])
                        print('continue failed split',row_p,val1[0][0])
                        continue

                    print("for lab this split is success",p)

                    if(p):


                        if(len(get_split)>1):
                            updated_day=p[0].split('--')
                            already_lab_day.append(updated_day[0])


                        update_ready.append(p)

                        collect_split.append(int(sp))

                        split_count+=1
                        break
                    
                    """
                    #success
                    temp=[]

                    temp=p

                    #update_ready=update_ready+p

                    #means this day and its periods are free 

                    if(temp):
                        split_count+=1
                        local_day=p
                        #update_ready=update_ready+p
                        
                    else:
                    #if(not update_ready):
                        print("Didn't find slot of lab")
                        break
                
                print("update_ready",update_ready)
                #start selecting another  faculty to reach faculty_count of this lab"""

                
                """if(update_ready):
                    updated_day=p[0].split('--')
                    already_lab_day.append(updated_day[0])"""



            print("")
            print("split_count",split_count,"get_split",len(get_split))

            if(split_count==len(get_split)):

                for pr in update_ready:
                    updated_day=pr[0].split('--')
                    already_lab_day.append(updated_day[0])


            else:

                #go for next lab
                print("lab:",row.subject,"didn't get free slots")
                continue#go to next lab

            #regular_list = [[1, 2, 3, 4], [5, 6, 7], [8, 9]]

            update_ready = [item for sublist in update_ready for item in sublist]

            if(flag_overflow):
                continue#go to next lab

            flag_fac_count=1 

            if(fac_lab_list):
                # head faculty is there get another faculty data according to list of update_ready list

                #designation_list=["proffesor","Assoc_prof","Asst_prof","Asst_prof_c"]

                designation_list=["Professor","Associate Professor","Assistant Professor","Assistant Professor(C)"]



                print("designation list",designation_list)

                head_fac_dept=Teacheradd.query.filter_by(name=fac_lab_list[0]).first()


                print("id:",head_fac_dept.id)
                position=designation_list.index(head_fac_dept.role)
                print("position:",position)

                course=row.classname.split("@")

                subject_dept=Subject.query.filter_by(name=row.subject,course=course[0]).first()

                print("subject_dept",subject_dept.dept,subject_dept.id,subject_dept.name)


                mycursor.execute("SELECT * from facultyadd WHERE branch='{0}' AND name <> '{1}' ORDER BY work_load ASC ".format(subject_dept.dept,fac_lab_list[0]))
             
                dep_fac=mycursor.fetchall()

                print("faculty lab count",row.f_count)

                #backup_list=[]
                
                for dep in dep_fac:

                    support_fac=Teacheradd.query.filter_by(name=dep[1]).first()

                    if(support_fac.exclude==1):
                        continue

                    mycursor.execute("SELECT total_Periods,allocated FROM classconfig_table WHERE faculty='{0}' AND type='{2}'".format(dep[1],row.classname,"THEORY"))

                    summ_check=mycursor.fetchall()

                    summ=0
                    for pr in summ_check:

                        print(p)
                        #this summm must to be scheduled
                        summ+=int(pr[0])-int(pr[1])

                    designation_dic={'Professor':'proffesor','Associate Professor':'Assoc_prof','Assistant Professor':'Asst_prof','Assistant Professor(C)':'Asst_prof_c'}

                    mycursor.execute("select `{0}` FROM day_periods WHERE id={1}".format(designation_dic[support_fac.role],1))

                    designation=mycursor.fetchall()

                    #day_period.query.get(1)

                    
                    if(support_fac.work_load+int(sp)+summ<=int(designation[0][0]) or (support_fac.role=="Assistant Professor(C)")):


                        print(dep[1],support_fac.work_load+int(sp)+summ," is less  his/her workload can procced")

                        #fac_lab_list.append(val[0][0])
                    else:
                        #backup_list.append(dep[1])
                        print(val[0][0],get_head.work_load+int(sp)+summ," is greater than his/her workload cannot  procced")
                        #break         
                        continue        
                        
                    while(len(fac_lab_list)<row.f_count):
                        select_flag=0

                        if(dep[3]=="Assistant Professor(C)"):
                            

                            #select  table and append to fac_lab_list

                            for check in update_ready:

                                print("in check for",check) 

                                period_split=check.split("--")
                                #select class
                                mycursor.execute("SELECT `{1}` FROM `{0}` WHERE DAY='{2}';".format(dep[1],period_split[1],period_split[0]))

                                val1=mycursor.fetchall()

                                if(val1[0][0]!="--"):
                                    select_flag=1
                                    break 

                            #if not free break
                            if(select_flag==0):
                                fac_lab_list.append(dep[1])
                                break
                        else:
                            new_position=designation_list.index(dep[3])

                            if(new_position>position):

                                #select  table and append to fac_lab_list

                                for check in update_ready:

                                    print("in check for",check) 

                                    period_split=check.split("--")
                                    #select class
                                    mycursor.execute("SELECT `{1}` FROM `{0}` WHERE DAY='{2}';".format(dep[1],period_split[1],period_split[0]))

                                    val1=mycursor.fetchall()

                                    if(val1[0][0]!="--"):
                                        select_flag=1
                                        break                                     
                                    

                                #if not free break

                                if(select_flag==0):
                                    position=new_position

                                    fac_lab_list.append(dep[1])
                                break
                            else:
                                break
                        

                print("final faculty list to this lab",fac_lab_list)
                    
                #print(dep[0][0],dep[0][1],dep[0][2],dep[0][3],dep[0][4])

                #k=0
                if(len(fac_lab_list)!=row.f_count):

                    """while(len(fac_lab_list)<row.f_count):

                        #fac_lab_list.append(backup_list[k])
                        k+=1"""
                    flag_fac_count=0

                print("")
            else:
                position=-1
                #head faculty is buzy assign random faculties according to list of update_ready list

                #designation_list=["proffesor","Assoc_prof","Asst_prof","Asst_prof_c"]

                designation_list=["Professor","Associate Professor","Assistant Professor","Assistant Professor(C)"]
                print("designation list",designation_list)

                #head_fac_dept=Teacheradd.query.filter_by(name=fac_lab_list[0]).first()


                #print("id:",head_fac_dept.id)
                #position=designation_list.index(head_fac_dept.role)
                #print("position:",position)

                course=row.classname.split("@")

                subject_dept=Subject.query.filter_by(name=row.subject,course=course[0]).first()

                print("subject_dept",subject_dept.dept,subject_dept.id,subject_dept.name)


                mycursor.execute("SELECT * from facultyadd WHERE branch='{0}'  ORDER BY work_load ASC ".format(subject_dept.dept))

                
                dep_fac=mycursor.fetchall()



                print("faculty lab count",row.f_count)

                #backup_list=[]
                for dep in dep_fac:

                    support_fac=Teacheradd.query.filter_by(name=dep[1]).first()

                    if(support_fac.exclude==1):
                        continue

                    mycursor.execute("SELECT total_Periods,allocated FROM classconfig_table WHERE faculty='{0}' AND type='{2}'".format(dep[1],row.classname,"THEORY"))

                    summ_check=mycursor.fetchall()

                    summ=0
                    for pr in summ_check:

                        #print(p)
                        #this summm must to be scheduled
                        summ+=int(pr[0])-int(pr[1])

                    designation_dic={'Professor':'proffesor','Associate Professor':'Assoc_prof','Assistant Professor':'Asst_prof','Assistant Professor(C)':'Asst_prof_c'}

                    mycursor.execute("select `{0}` FROM day_periods WHERE id={1}".format(designation_dic[support_fac.role],1))

                    designation=mycursor.fetchall()

                    #day_period.query.get(1)

                    
                    if(support_fac.work_load+int(sp)+summ<=int(designation[0][0]) or (support_fac.role=="Assistant Professor(C)") ):


                        print(dep[1],support_fac.work_load+int(sp)+summ," is less than his/her workload can procced")

                        #fac_lab_list.append(val[0][0])
                    else:
                        print(dep[1],support_fac.work_load+int(sp)+summ," is greater than his/her workload cannot  procced")
                        #break       
                        continue          

                
                    while(len(fac_lab_list)<row.f_count):
                        select_flag=0

                        if(dep[3]=="Assistant Professor(C)"):

                            #select  table and append to fac_lab_list

                            for check in update_ready:

                                print("in check for",check)

                                period_split=check.split("--")
                                #select class
                                mycursor.execute("SELECT `{1}` FROM `{0}` WHERE DAY='{2}';".format(dep[1],period_split[1],period_split[0]))

                                val1=mycursor.fetchall()

                                if(val1[0][0]!="--"):
                                    select_flag=1
                                    break 

                            #if not free break
                            if(select_flag==0):
                                fac_lab_list.append(dep[1])
                                position=3
                            break
                        else:

                            #backup_list.append(dep[1])
                        
                            new_position=designation_list.index(dep[3])

                            if(new_position>position):

                                #select  table and append to fac_lab_list

                                for check in update_ready:

                                    period_split=check.split("--")
                                    #select class
                                    mycursor.execute("SELECT `{1}` FROM `{0}` WHERE DAY='{2}';".format(dep[1],period_split[1],period_split[0]))

                                    val1=mycursor.fetchall()

                                    if(val1[0][0]!="--"):
                                        select_flag=1

                                        break                                     
                                    

                                #if not free break
                                 #if not free break
                                if(select_flag==0):
                                    position=new_position

                                    fac_lab_list.append(dep[1])
                                break
                                
                            else:
                                break


                #k=0
                if(len(fac_lab_list)!=row.f_count):


                    flag_fac_count=0


                print("final faculty list to this lab",fac_lab_list)
                    
                #print(dep[0][0],dep[0][1],dep[0][2],dep[0][3],dep[0][4])

            if(flag_fac_count==1):#update it 

                #updation code 

                print(row.subject,"faculty count satisfied ")


                #update split by appending '/' and scale

                data=classconfig_table.query.get(int(row.id))
    


                data.split=data.split+"/"+data.split
                db.session.commit()

                data.faculty=(',').join(fac_lab_list)
                db.session.commit()


                for u in update_ready:

                    
                    period_split=u.split("--")

                    #update class
                    mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE Day='{2}';".format(period_split[1],row.classname,period_split[0],row.subject))
                    mydb.commit()

                    #select fac

                    for fac_u in fac_lab_list:
    
                        #update faculty remove "--"
                        mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE `Day`='{2}';".format(period_split[1],fac_u,period_split[0],row.classname+"/"+row.subject))

                        mydb.commit()

                        faculty=Teacheradd.query.filter_by(name=fac_u).first()

                        faculty.work_load=faculty.work_load+1
                        db.session.commit()                                    


                    #update workload,split,allocated
                    all_lab=row.lab.split(',')
                    for lab_u in all_lab:

                        mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE `Day`='{2}';".format(period_split[1],lab_u,period_split[0],row.classname+"/"+row.subject))

                        mydb.commit()



                    data=classconfig_table.query.get(int(row.id))
                    
                    data.allocated=data.allocated+1
                    db.session.commit()

                    if(data.dayperiod=="@"):
                        data.dayperiod=u
                        db.session.commit()
                    else:
                        data.dayperiod= data.dayperiod +","+u
                        db.session.commit()
                    

                print(already_lab_day)
                print(fac_lab_list)
                print(update_ready)

                updated_day=p[0].split('--')

                already_lab_day.append(updated_day[0])

                print("")
            else:
                #remove those recently added elements
                print("faculty count for this is doesnot satisfied")
                k=0
                while(k<len(get_split)):
                    already_lab_day.pop()
                    k=k+1

            update_ready=[]
                        


        print("classname:",get_data['data'])





        mycursor = mydb.cursor(buffered=True)
        #generate electives


        #first collect the elective which are incomplete

        #all_elec_class=classconfig_table.query.filter(type.like('ELECTIVE/%') ).all()

        mycursor.execute("SELECT DISTINCT classname FROM classconfig_table WHERE type LIKE 'ELECTIVE/%' AND total_periods <> allocated ")

        val=mycursor.fetchall()

        all_class=[]
        all_elec=[]

        print("ELECTIVES containing classes")
        print(val)
        for i in val:
            all_class.append(i[0])
            
            print(i[0])

        mycursor.execute("SELECT DISTINCT type,split  FROM classconfig_table WHERE type LIKE 'ELECTIVE/%' AND total_periods <> allocated")

        val=mycursor.fetchall()

        print("Number of Electives")
        for i in val:
            all_elec.append([i[0],i[1]])
            print(i[0],i[1])

        already_day=[]

        mycursor.execute("SELECT DISTINCT dayperiod FROM classconfig_table WHERE type LIKE 'ELECTIVE/%' AND dayperiod <>'@' ")

        day_elec=mycursor.fetchall()

        if(day_elec):

            for day_elec_row in day_elec:

                print("already elective day",day_elec_row)

                if(day_elec_row[0] in already_day):

                    already_day.append(day_elec_row[0])




        for i in all_elec:

            print("generating",i[0])
            #assuming all subjects having same subject split of particular elective

            split_slash=i[1].split('/')

            li=[]

            print(split_slash)
            
            if(len(split_slash)>1):
                
                for p in split_slash:
                        li.append(p.split(','))

                for j in li[1]:
                        if(j in li[0]):
                                    li[0].remove(j)

                li=li[0]

                print(li)
            else:

                li=split_slash[0].split(',')

            print(i[0],":",li) 

            li.sort(reverse=True)
            for sp in li:
                #already_day=[]
                #take random extreme start of day mrng or afternoon

                get_day=day_period.query.get(1)
                day=int(get_day.day)

                print("max day:",day)

                day_list=['MON','TUE','WED','THU','FRI','SAT','SUN']  

                #while(len(already_day)<=2):
                #while(len(already_day)<=len(day_list[:day])):

                day_list=day_list[:day]

                random.shuffle(day_list)

                
                
                for random_day in day_list:      
                    #random_day= random.choice()

                    if(random_day in already_day):
                        continue

                    print("random_day",random_day)

                    already_random_ext=[]

                    update_status=1

                    while(len(already_random_ext)!=2 ):
                        #0->morning extreme
                        #1-afternoon extreme
                        random_ext=random.choice([0,1])

                        if(random_ext  in already_random_ext):

                            continue

                        if(random_ext==0):#morning
                            #generate strings
                            start=1
                            print("morning selected start:",start)
                        else:#afternoon
                            start=int(get_day.morning_periods)+1
                            print("afternoon selected start:",start)

                        #generate strings
                        period_collection=[]

                        for start_scale in range(start,start+int(sp)):

                            #collect data:
                            period_collection.append(random_day+"--"+str(start_scale))


                        print("period collection:",period_collection)

                        #collect class data and faculty data and check on these periods collection

                        #get_elec_data=classconfig_table.query.filter_by(type=i[0],total_periods!=allocated).all()


                        all_fac_class=[]

                        #mycursor.execute("SELECT DISTINCT classname,type FROM classconfig_table WHERE total_periods <> allocated ")


                        mycursor.execute("SELECT * FROM classconfig_table WHERE total_periods <> allocated ")

                        get_elec_data=mycursor.fetchall()

                        all_class_ele=[]
                        all_fac_ele=[]

                        print("list of classes:")
                        for row in get_elec_data:

                            if(row[2]==i[0]):
                                print(row[4])
                                all_class_ele.append(row[4])

                        #mycursor.execute("SELECT DISTINCT faculty,type FROM classconfig_table WHERE total_periods <> allocated ")

                        #get_elec_data=mycursor.fetchall()

                        final_all=[]

                        print("list of faculties:")
                        for row in get_elec_data:

                            if(row[2]==i[0]):
                                print(row[3])
                                all_fac_ele.append(row[3])

                                final_all.append(row)

                        #if all are free at these slots assign them and go for next split

                        #else and period already_random_ext.append(random_ext)

                        all_fac_class=all_fac_ele+all_class_ele
                        #print("final_all",final_all)
                        #print("\n\nfinal list:",all_fac_class)

                        for p in final_all:
                            print(p[0])

                        status=0
                        print("period_collection",period_collection)
                        for final in all_fac_class:

                            for period in period_collection:

                                period_split=period.split("--")

                                print("period_split",period_split)

                                mycursor.execute("SELECT `{1}` FROM `{0}` WHERE DAY='{2}';".format(final,period_split[1],period_split[0]))

                                val1=mycursor.fetchall()

                                if(val1[0][0]=="--"):
                                    print("ok",final,period)
                                else:

                                    status=1
                                    print("plz select alternate session of this same day if not tried")
                                    break

                            if(status==1):#breaking outer forloop control  going to after outer for loop and control goes to down the outer while
                                already_random_ext.append(random_ext)
                                break

                                #continue

                        #start update throught respective classes/faculties
                        if(status==0):
                            print("control came to update part")

                            
                            for final in final_all:

                                for period in period_collection:

                                    period_split=period.split("--")

                                    #select class
                                    mycursor.execute("SELECT `{1}` FROM `{0}` WHERE DAY='{2}';".format(final[4],period_split[1],period_split[0]))

                                    val=mycursor.fetchall()

                                    if(val1[0][0]=="--"):#update it
                                        #update class
                                        mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE Day='{2}';".format(period_split[1],final[4],period_split[0],i[0]))
                                        mydb.commit()

                                    #select fac
                                    mycursor.execute("SELECT `{1}` FROM `{0}` WHERE DAY='{2}';".format(final[3],period_split[1],period_split[0]))

                                    val=mycursor.fetchall()

                                    if(val[0][0]=="--"):
                                    #update faculty remove "--"
                                        mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE `Day`='{2}';".format(period_split[1],final[3],period_split[0],final[4]+'/'+final[1]))

                                        mydb.commit()
                                        faculty=Teacheradd.query.filter_by(name=final[3]).first()

                                        faculty.work_load=faculty.work_load+1
                                        db.session.commit()                                    
                                    else:
                                        mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE `Day`='{2}';".format(period_split[1],final[3],period_split[0],final[4]+'/'+val[0][0]))

                                        mydb.commit()

                                    #update workload,split,allocated



                                    data=classconfig_table.query.get(int(final[0]))
                                    
                                    data.allocated=data.allocated+1
                                    db.session.commit()

                                    if(data.dayperiod=="@"):
                                        data.dayperiod=period
                                        db.session.commit()
                                    else:
                                        data.dayperiod= data.dayperiod +","+period
                                        db.session.commit()



                                #update split by appending '/' and scale
                                scale=int(sp)
                                if(data.allocated-scale==0):
                                    #append "/"

                                    data.split=data.split+"/"+str(scale)
                                    db.session.commit()
                                else:
                                    data.split=data.split+","+str(scale)
                                    db.session.commit()

                                
            
                            already_day.append(random_day)
                            update_status=0
                            break#while(len(already_day))

                            #update successfull u need to got next split

                            #print(exit)
                            #exit()

                            #already_random_ext.append(random_ext)        
                            #break    
                        else:#control is going to while(len(already_random_ext)!=2)

                            print("session failed {0} day select another extreme of this day",format(period))
                            #while 
                            
                            #go to next split of this elective
                            #break
                    #break will further control to next split
                    #while(len(already_random_ext)==2) end

                    if(update_status==0):#success

                        break #break the loop

                    if(len(already_random_ext)==2 ):

                        #go to next random day
                        print("day failed") 
                        already_day.append(random_day)
                        print("length of already day",len(already_day))
                        #while                 
                    

                if(len(already_day)==len(day_list[:day]) and update_status==1):

                    return jsonify({'msg':"{0} does not fitted into point of extreme start(morning or afternoon) of day ".format(i[0])})


        #generating theory


        mycursor = mydb.cursor(buffered=True)
        print("")
        print("")
        print("")
        print("")
        print("theory generating")

        get_class_theory=classconfig_table.query.filter_by(classname=get_data['data'],type="THEORY").all() 

        already_theory_day=[]#on successful updation of single day add day to this of split

        for a in get_class_theory:


            if(a.allocated==a.total_periods):

                already_theory_day=[]
                continue
                #already fully assigned

            print(a.subject)

            if(a.allocated!=a.total_periods and a.dayperiod!='@'):

                #collect already assigned day

                day_split=a.dayperiod.split(',')

                for day_index in day_split:

                    only_day=day_index.split('--')

                    if(only_day not in already_theory_day):

                        already_theory_day.append(only_day[0])    



            split_slash=a.split.split('/')

            li=[]

            print("split_slash thoery",split_slash)
            
            if(len(split_slash)>1):
                
                for p in split_slash:
                        li.append(p.split(','))

                for j in li[1]:
                        if(j in li[0]):
                                    li[0].remove(j)

                li=li[0]

                print(li)
            else:

                li=split_slash[0].split(',')

            #print(i[0],":",li)
                
            #ready start here for one theory subject
            print(li)
            li.sort(reverse=True)
            for sp in li:


                #collecting days

                get_day=day_period.query.get(1)
                day=int(get_day.day)

                print("max day:",day)

                day_list=['MON','TUE','WED','THU','FRI','SAT','SUN']  

                day_list=day_list[:day] 

                random.shuffle(day_list)
                
                period_colloction_final=[]

                start1=1

                end1=int(get_day.morning_periods)

                if(int(get_day.morning_periods)+1 < int(get_day.periods)):

                    start2=int(get_day.morning_periods)+1

                    end2=int(get_day.periods)
                else:
                    start2=1
                    end2=int(get_day.periods)

                print("start1,end1,start2,end2",start1,end1,start2,end2)

                period_colloction_final=[]

                for d in day_list:
                                    
                    period_collection=[]

                    
                    if(d in already_theory_day):
                        continue

                    #if sp is even periods(2,4) should start from 1 3 5 
                    #if sp is odd(1,3,5) periods should start from 1 3 5

                    for one in range(start1,end2-int(sp)+2):

                        period_collection=[]
                        for index in range(int(sp)):

                            #print(d,"--",one+index)

                            st=d+'--'+str(int(one+index))

                            period_collection.append(st)

                        print(period_collection)

                        fir_arr=period_collection[0].split('--')

                        sec_arr=period_collection[-1].split('--')


                        if( (((int(fir_arr[-1])==start1) or (int(sec_arr[-1])==end1)) or((int(fir_arr[-1])==start2) or (int(sec_arr[-1])==end2))) and ((int(fir_arr[-1])!=end1) and (int(sec_arr[-1])!=start2) ) ):


                                period_colloction_final.append(period_collection)

                if(period_colloction_final):
                    print("period collection_final:",period_colloction_final)
                else:
                    flag_overflow=1
                    print("do it manually no slot found")
                    break
                
                
                final_collection=[]
                #got possible slots of theory
                flag_update=1
                for one_row in period_colloction_final:

                    for one in one_row:

                        period_split=one.split('--')

                        #select faculty

                        if(int(period_split[1])==int(start1) or int(period_split[1])==int(start2) ):
                        
                            mycursor.execute("SELECT `{0}` FROM `{1}` WHERE DAY='{2}';".format(period_split[1],a.faculty,period_split[0]))

                            val=mycursor.fetchall()
                            print(a.faculty,val[0][0])

                            if(val[0][0]!='--'):
                                flag_update=0
                                print("if",a.faculty,val[0][0])
                                break

                            mycursor.execute("SELECT `{0}` FROM `{1}` WHERE DAY='{2}';".format(period_split[1],a.classname,period_split[0]))

                            cls=mycursor.fetchall()
                            if(cls[0][0]!='--'):
                                flag_update=0
                                print("if",a.classname,cls[0][0])
                                break 

                        else:

                            before=str(int(period_split[1])-1)

                            mycursor.execute("SELECT `{0}` FROM `{1}` WHERE DAY='{2}';".format(before,a.faculty,period_split[0]))     

                            before=mycursor.fetchall()

                            if(before[0][0]!='--'):
                                flag_update=0
                                print("else",a.faculty,before[0][0])
                                break

            

                            mycursor.execute("SELECT `{0}` FROM `{1}` WHERE DAY='{2}';".format(period_split[1],a.faculty,period_split[0]))

                            fac=mycursor.fetchall()
                            print(a.faculty,val[0][0])

                            if(fac[0][0]!='--'):
                                flag_update=0
                                print(a.faculty,fac[0][0])
                                break    

                            
                            mycursor.execute("SELECT `{0}` FROM `{1}` WHERE DAY='{2}';".format(period_split[1],a.classname,period_split[0]))

                            cls=mycursor.fetchall()
                            if(cls[0][0]!='--'):
                                flag_update=0
                                print(a.classname,cls[0][0])
                                break 
                                        

                        #all ok that means 

                        final_collection=one_row

                        for u in final_collection:
                        #start updating 

                            period_split=u.split('--')

                            #faculty update

                            mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE `Day`='{2}';".format(period_split[1],a.faculty,period_split[0],a.classname+'/'+a.subject))                    

                            mydb.commit()
                                        
                            faculty=Teacheradd.query.filter_by(name=a.faculty).first()

                            faculty.work_load+=1
                            db.session.commit()

                            a.allocated=a.allocated+1
                            db.session.commit()

                            if(a.dayperiod=="@"):
                                a.dayperiod=u
                                db.session.commit()
                            else:
                                a.dayperiod= a.dayperiod +","+u
                                db.session.commit()  

                            #update class                  
                            mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE `Day`='{2}';".format(period_split[1],a.classname,period_split[0],a.subject))                    

                            mydb.commit()
                            flag_update=1

                        break
                    
                    #update split by appending '/' and scale

                    if(flag_update==1):

                        if('/' in a.split):

                            a.split=a.split+","+str(sp)
                            db.session.commit()
                        else:
                            a.split=a.split+"/"+str(sp)
                            db.session.commit()

                        already_theory_day.append(period_split[0])
                        break
                    else:
                        already_theory_day=[]


                #break #break outer loop


            already_theory_day=[]
            #deallocate already


            print("")
            print("")



            #last you have to clear

        #generating SRP


        mycursor = mydb.cursor(buffered=True)
        print("")
        print("")
        print("")
        print("")
        print("SRP generating")

        get_class_theory=classconfig_table.query.filter_by(classname=get_data['data'],type="PROJECT").all() 

        already_theory_day=[]#on successful updation of single day add day to this of split

        for a in get_class_theory:


            if(a.allocated==a.total_periods):

                already_theory_day=[]
                continue
                #already fully assigned

            if(a.faculty!="NA"):

                msg+=" ,{0} subject is already been assigned with faculty eiether you remove the faculty or do manually".format(a.subject)
                continue
            print(a.subject)

            if(a.allocated!=a.total_periods and a.dayperiod!='@'):

                #collect already assigned day

                day_split=a.dayperiod.split(',')

                print("already half assigned",day_split)

                for day_index in day_split:

                    only_day=day_index.split('--')

                    if(only_day not in already_theory_day):

                        already_theory_day.append(only_day[0])    



            split_slash=a.split.split('/')

            li=[]

            print("split_slash thoery",split_slash)
            
            if(len(split_slash)>1):
                
                for p in split_slash:
                        li.append(p.split(','))

                for j in li[1]:
                        if(j in li[0]):
                                    li[0].remove(j)

                li=li[0]

                print(li)
            else:

                li=split_slash[0].split(',')

            #print(i[0],":",li)
                
            #ready start here for one theory subject
            print(li)
            li.sort(reverse=True)

            fac_srp_list=[]


            for sp in li:

                final_update=0
                #collecting days

                get_day=day_period.query.get(1)
                day=int(get_day.day)

                print("max day:",day)

                day_list=['MON','TUE','WED','THU','FRI','SAT','SUN']  

                day_list= day_list[:day]

                random.shuffle(day_list)
                
                period_colloction_final=[]

                start1=1

                end1=int(get_day.morning_periods)

                if(int(get_day.morning_periods)+1 < int(get_day.periods)):

                    start2=int(get_day.morning_periods)+1

                    end2=int(get_day.periods)
                else:
                    start2=1
                    end2=int(get_day.periods)


                period_colloction_final=[]

                for d in day_list:
                                    
                    period_collection=[]

                    
                    if(d in already_theory_day):
                        continue

                    #if sp is even periods(2,4) should start from 1 3 5 
                    #if sp is odd(1,3,5) periods should start from 1 3 5

                    for one in range(start1,end2-int(sp)+2):

                        period_collection=[]
                        for index in range(int(sp)):

                            #print(d,"--",one+index)

                            st=d+'--'+str(int(one+index))

                            period_collection.append(st)

                        print(period_collection)

                        fir_arr=period_collection[0].split('--')

                        sec_arr=period_collection[-1].split('--')


                        if( (((int(fir_arr[-1])==start1) or (int(sec_arr[-1])==end1)) or((int(fir_arr[-1])==start2) or (int(sec_arr[-1])==end2))) and ((int(fir_arr[-1])!=end1) and (int(sec_arr[-1])!=start2) ) ):
                                period_colloction_final.append(period_collection)

                if(period_colloction_final):
                    print("period collection_final:",period_colloction_final)
                else:
                    flag_overflow=1
                    print("do it manually no slot found")
                    break
                
                
                final_collection=[]
                #got possible slots of theory
                flag_update=1

                #LOAD RANDOM FACULTY ACCORDING TO THEIR WORKLOADS

                course=a.classname.split("@")

                subject_dept=Subject.query.filter_by(name=a.subject,course=course[0]).first()

                print("subject_dept",subject_dept.dept,subject_dept.id,subject_dept.name)


                mycursor.execute("SELECT * from facultyadd WHERE branch='{0}'  ORDER BY work_load ASC ".format(subject_dept.dept))

                
                dep_fac=mycursor.fetchall()



                print("faculty SRP count",a.f_count)   

                for dep in dep_fac:

                    print("srp faculty",dep[1])

                    support_fac=Teacheradd.query.filter_by(name=dep[1]).first()
                    
                    if(support_fac.exclude==1):
                        continue

                    mycursor.execute("SELECT total_Periods,allocated FROM classconfig_table WHERE faculty='{0}' AND type='{1}'".format(dep[1],"THEORY"))

                    summ_check=mycursor.fetchall()

                    summ=0
                    for pr in summ_check:

                        #print("p",p)
                        #this summm must to be scheduled
                        summ+=int(pr[0])-int(pr[1])

                    designation_dic={'Professor':'proffesor','Associate Professor':'Assoc_prof','Assistant Professor':'Asst_prof','Assistant Professor(C)':'Asst_prof_c'}

                    mycursor.execute("select `{0}` FROM day_periods WHERE id={1}".format(designation_dic[support_fac.role],1))

                    designation=mycursor.fetchall()

                    #day_period.query.get(1)

                    
                    if(support_fac.work_load+int(sp)+summ<=int(designation[0][0]) or (support_fac.role=="Assistant Professor(C)") ):


                        print(dep[1],support_fac.work_load+int(sp)+summ," is less than his/her workload can procced")

                        #append it 

                        fac_srp_list.append(dep[1])

                    else:
                        print(dep[1],support_fac.work_load+int(sp)+summ," is greater than his/her workload cannot  procced")
                        continue

                if(len(fac_srp_list)<a.f_count):

                    print("{0} is not having enough Faculty count".format(a.subject))
                    break

                    #return jsonify({'msg':"{0} is not having enough Faculty count".format(a.subject)})

                #all fac of length found and make update 

                print(fac_srp_list)
                for fac in fac_srp_list:
                    print("faculty name:",fac)

                print("faculty printing completed")

                print(period_colloction_final)
                for one_row in period_colloction_final:           

                    flag_update=1

                    print("inside for split",one_row)

                    for one in one_row:

                        print("inside period:",one)
                        period_split=one.split('--')

                        mycursor.execute("SELECT `{0}` FROM `{1}` WHERE DAY='{2}';".format(period_split[1],a.classname,period_split[0]))

                        cls=mycursor.fetchall()
                        print(cls)

                        if(cls[0][0]!='--'):
                            flag_update=0
                            #print("if",a.classname,cls[0][0])
                            break 

                    if(flag_update==1):
                        print("free srp slot found",one_row)

                        final_fac_list=[]
                        for fac in fac_srp_list:
                            fac_update=1
                            if(len(final_fac_list)==a.f_count):
                                break
                            else:
                                for one in one_row:

                                    period_split=one.split('--')

                                    if(int(period_split[1])==start1 or int(period_split[1])==start2 ):

                                        mycursor.execute("SELECT `{0}` FROM `{1}` WHERE DAY='{2}';".format(period_split[1],fac,period_split[0]))

                                        cls=mycursor.fetchall()

                                        if(cls[0][0]!='--'):
                                            fac_update=0
                                            #print("if",a.classname,cls[0][0])
                                            break
                                    else:
                                        before=str(int(period_split[1])-1)

                                        mycursor.execute("SELECT `{0}` FROM `{1}` WHERE DAY='{2}';".format(before,fac,period_split[0]))

                                        cls=mycursor.fetchall()

                                        if(cls[0][0]!='--'):
                                            fac_update=0
                                            #print("if",a.classname,cls[0][0])
                                            break 


                                        mycursor.execute("SELECT `{0}` FROM `{1}` WHERE DAY='{2}';".format(period_split[1],fac,period_split[0]))

                                        cls=mycursor.fetchall()

                                        if(cls[0][0]!='--'):
                                            fac_update=0
                                            #print("if",a.classname,cls[0][0])
                                            break                                                                    

                                if(fac_update==1):
                                    final_fac_list.append(fac)

                        print("final faculty for srp is",final_fac_list)

                        if(len(final_fac_list)==a.f_count):

                                #update is start here 

                                for one in one_row:

                                    period_split=one.split('--')

                                    #update class                  
                                    mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE `Day`='{2}';".format(period_split[1],a.classname,period_split[0],a.subject))                    

                                    mydb.commit()  


                                    a.allocated=a.allocated+1
                                    db.session.commit()

                                    if(a.dayperiod=="@"):
                                        a.dayperiod=one
                                        db.session.commit()
                                    else:
                                        a.dayperiod= a.dayperiod +","+one
                                        db.session.commit() 

                                    #update faculty
                                    for inner in final_fac_list:

                                        mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE `Day`='{2}';".format(period_split[1],inner,period_split[0],a.classname+'/'+a.subject))                    

                                        mydb.commit()

                                        faculty=Teacheradd.query.filter_by(name=inner).first()

                                        faculty.work_load+=1
                                        db.session.commit()  

                                        final_update=1


                                if(final_update==1): 

                                    if('/' in a.split):

                                        a.split=a.split+","+str(sp)
                                        db.session.commit()
                                    else:
                                        a.split=a.split+"/"+str(sp)
                                        db.session.commit()
                                    
                                    a.faculty=(',').join(final_fac_list)
                                    db.session.commit()

                                already_theory_day.append(period_split[0])
                                break                            

                        else:

                            print("faculty count does fitted any free slots")

                   

                        #we will take this and update it
                        break 


            #last you have to clear


        mycursor.execute("SELECT * FROM `{0}`;".format(get_data['data']))
        row=mycursor.fetchall()

        all_data=classconfig_table.query.all()

        get_period=day_period.query.get(1)

        return jsonify( {'data': render_template('classtable.html',all_data=row,periods=get_period.periods),'describe':render_template('classtable_describe.html',classname=get_data['data'],all_data=all_data),'msg':msg}) 

    else:

        return redirect(url_for('dashboard'))


    #return jsonify({'data':"method passed"})

@app.route('/excel_report', methods = ['GET', 'POST'])
def excel_report():

    if 'loggedin' in session:
    
        file=excel_method()
        #path="E:/project-old/"

        #flag=1

        if(file):

            return jsonify({'msg':"excel report success"})
        else:
            return jsonify({'msg':"excel report fail"})
    else:

        return redirect(url_for('dashboard'))

def excel_method():

    if 'loggedin' in session:


        mycursor = mydb.cursor(buffered=True) 
        
        wb = Workbook()


        all_branch=Teacheradd.query.with_entities(Teacheradd.branch).distinct()

        

        count=0
        for branch in all_branch:

            print("Branch Faculty:",branch.branch)

            ws1 = wb.create_sheet("Sheet_A",count)

            ws1.title = branch.branch
            row=col=1

            all_teachers=Teacheradd.query.filter_by(branch=branch.branch).all()

            for p in all_teachers:
                print(p.name)

            for teacher in all_teachers:

                if(show_all_tables((teacher.name).lower())):

                    print("printing {0} table",format(teacher.name))

                    #print name of the faculty

                    ws1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+5)

                    ws1.cell(row=row,column=col,value="Faculty Name:"+teacher.name).font = Font(size = 14, bold = True)

                    thin_border = Border(outline=Side(style='thin'))
                    ws1.cell(row=row,column=col).border=thin_border

                    ws1.row_dimensions[row].height = 20
                    #st=chr(col+64)
                    #print("ascii",st)
                    #ws1.column_dimensions[st].width = 30.0

                    row+=1

                    #start printing faculty table
                    #cols=fields,Type Null Key Default Extra


                    try:

                        mycursor.execute("SHOW COLUMNS FROM  `{0}`".format(teacher.name))

                    except:

                        print("An exception occurred")
                        row+=2
                        col=1 
                        continue

                    val=mycursor.fetchall()

                    for j in val:
                        ws1.cell(row=row,column=col,value=j[0]).font = Font(size = 12, bold = True)
                        
                        ws1.column_dimensions[chr(col+64)].width = 20.0
                        col+=1
                    ws1.row_dimensions[row].height = 50


                    row+=1
                    col=1
                    index=0
                    mycursor.execute("SELECT * FROM `{0}`".format(teacher.name))

                    val=mycursor.fetchall()

                    for j in val:

                        print("row:",j)
                        col=1
                        for r in j:
                            #print("q:",q)

                            #print(r[q])

                            ws1.cell(row=row,column=col,value=r)
                            ws1.column_dimensions[chr(col+64)].width = 20.0
                            col+=1

                        ws1.row_dimensions[row].height = 50 
                        row+=1
                row+=2
                col=1      

            count+=1

        file=datetime.datetime.now().strftime("Faculty-reports%Y%m%d%H%M%S")
        wb.save(file)
        wb.close()


        time.sleep(2.4)
        
        #redirect('/download/{0}'.format(file))
        
        #file1 = open("faculty-report.xlsx", "rb")

            
                
        return 1
    else:
        return redirect(url_for('dashboard'))

    
@app.route('/delete_all', methods = ['GET', 'POST'])

def delete_all():


    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table')) 

        mycursor = mydb.cursor(buffered=True)

        get_data=request.get_json()

        #class_reset



        flag=class_reset(get_data['data'])

        if(flag):  
            print("get_data",get_data)

            mycursor.execute("SELECT * FROM `{0}`;".format(get_data['data']))
            row=mycursor.fetchall()
            all_data=classconfig_table.query.all()

            get_period=day_period.query.get(1)

            return jsonify( {'data': render_template('classtable.html',all_data=row,periods=get_period.periods),'describe':render_template('classtable_describe.html',classname=get_data['data'],all_data=all_data)})


            return jsonify( {'data':"success"})
    else:
        return redirect(url_for('dashboard'))



def class_reset(get_class):

    all_class_data=classconfig_table.query.filter_by(classname=get_class).all()


    if(all_class_data):
        for i in all_class_data:
            print(i.subject)

            if(i.type=="THEORY"):

                if(i.dayperiod !="@"):

                    get_day_period=i.dayperiod

                    print(get_day_period)

                    get_faculty=i.faculty
                    work=Teacheradd.query.filter_by(name=get_faculty).first()
                    arr=get_day_period.split(',')
                    print("arr",arr)
                    
                    for j in arr:
                        day_period_split=j.split('--')
                        day=day_period_split[0]
                        period=day_period_split[1]

                        mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(get_faculty,period,day))
                        mydb.commit()

                        mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(get_class,period,day))
                        mydb.commit()

                        work.work_load-=1

                        db.session.commit()

                    i.allocated=0
                    db.session.commit()

                    i.dayperiod='@'
                    db.session.commit() 

                    spl=(i.split).split('/')
                    i.split=spl[0]
                    db.session.commit()

                else:
                    print("else already @",i.subject)                  

                    
                        
            elif(i.type=="LAB"):
                if(i.dayperiod !="@"):
                    get_day_period=i.dayperiod

                    print(get_day_period)

                    get_faculty_split=(i.faculty).split(",")

                    get_lab_split=(i.lab).split(",")


                    arr=get_day_period.split(',')

                    print("arr",arr)
                    
                    for j in arr:
                        day_period_split=j.split('--')
                        day=day_period_split[0]
                        period=day_period_split[1]

                        for fac in get_faculty_split:

                            mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(fac,period,day))
                            mydb.commit()

                            work=Teacheradd.query.filter_by(name=fac).first()

                            work.work_load-=1
                            db.session.commit()

                        for lab in get_lab_split:

                            mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(lab,period,day))
                            mydb.commit()

                        mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(get_class,period,day))
                        mydb.commit()        

                    i.allocated=0
                    db.session.commit()

                    i.dayperiod='@'
                    db.session.commit() 

                    first=(i.split).split('/')

                    i.split=first[0]
                    db.session.commit()

                    i.faculty='NA'
                    db.session.commit()

                else:
                    print("else lab already @",i.subject)  

            elif(i.type=='PROJECT'):
                if(i.dayperiod !="@"):
                    get_day_period=i.dayperiod

                    print(get_day_period)
                    get_faculty_split=(i.faculty).split(",")

                    arr=get_day_period.split(',')
                    print("arr",arr)
                    
                    for j in arr:
                        day_period_split=j.split('--')
                        day=day_period_split[0]
                        period=day_period_split[1]

                        for fac in get_faculty_split:

                            mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(fac,period,day))
                            mydb.commit()

                            work=Teacheradd.query.filter_by(name=fac).first()

                            work.work_load-=1
                            db.session.commit()

                        mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(get_class,period,day))
                        mydb.commit()        

                    i.allocated=0
                    db.session.commit()

                    i.dayperiod='@'
                    db.session.commit()   

                    first=(i.split).split('/')

                    i.split=first[0]
                    db.session.commit()         

                    i.faculty='NA'
                    db.session.commit()             
                else:
                    print("else project already @",i.subject)

            else:#elective
                if(i.dayperiod !="@"):

                    get_day_period=i.dayperiod

                    print(get_day_period)

                    get_faculty=i.faculty

                    work=Teacheradd.query.filter_by(name=get_faculty).first()

                    arr=get_day_period.split(',')

                    print("arr",arr)
                    
                    for j in arr:
                        day_period_split=j.split('--')
                        day=day_period_split[0]
                        period=day_period_split[1]

                        mycursor.execute("SELECT `{1}` FROM `{0}` WHERE DAY='{2}';".format(get_faculty,period,day))

                        #split selected by '/'

                        print("select complted")
                        val=mycursor.fetchall()

                        val=(val[0][0]).split('/')

                        print(val)

                        if(len(val)==2):

                            mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(get_faculty,period,day))
                            mydb.commit()

                            mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(get_class,period,day))
                            mydb.commit()

                            work.work_load-=1

                            db.session.commit()

                        else:
                            #just update faculty cell by removing current class and don't decrease the workload
                            #val.remove(get_data['data'])

                            val.remove(get_class)

                            delimeter="/"

                            print("fac_cell",val)

                            st=('/').join(val)

                            print(st)

                            mycursor.execute("UPDATE `{0}` SET `{1}`='{2}' WHERE `Day`='{3}';".format(get_faculty,period,st,day))
                            mydb.commit()

                            mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(get_class,period,day))
                            mydb.commit()

                            print("elective string:",st)


                    i.allocated=0
                    db.session.commit()

                    i.dayperiod='@'
                    db.session.commit()

                    first=(i.split).split('/')

                    i.split=first[0]
                    db.session.commit()

                else:
                    print("else elective already @",i.subject)    
   
        return 1

@app.route('/delete', methods = ['GET', 'POST'])

def delete():

    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table')) 
        mycursor = mydb.cursor(buffered=True)
        get_data=request.get_json()

        print(get_data)

        if(get_data['status']=="only_id" ):#delete entire subject in that class
            
            id_split=get_data['ids'].split(',')

            print(id_split)

            for i in id_split:
                flag=0
                ids=int(i)

                get_id_data=classconfig_table.query.get(ids)

                dayperiod=(get_id_data.dayperiod).split(",")

                print(get_id_data,dayperiod)

                for j in dayperiod:

                    if(j=="@"):
                        flag=1
                        break
                        #return jsonify( {'msg': "periods  not been assigned yet"})
                    

                    if(get_id_data.type=="THEORY"):

                        get_faculty=get_id_data.faculty
                        work=Teacheradd.query.filter_by(name=get_faculty).first()


                        day_period_split=j.split('--')
                        day=day_period_split[0]
                        period=day_period_split[1]      
                        
                        mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(get_faculty,period,day))
                        mydb.commit()   

                        mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(get_data['class'],period,day))
                        mydb.commit()   

                        work.work_load-=1

                        db.session.commit() 

                    elif(get_id_data.type=="LAB"):
                        get_faculty_split=(get_id_data.faculty).split(",")

                        get_lab_split=(get_id_data.lab).split(",")

                        day_period_split=j.split('--')
                        day=day_period_split[0]
                        period=day_period_split[1]

                        for fac in get_faculty_split:

                            mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(fac,period,day))
                            mydb.commit()

                            work=Teacheradd.query.filter_by(name=fac).first()

                            work.work_load-=1
                            db.session.commit()

                        for lab in get_lab_split:

                            mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(lab,period,day))
                            mydb.commit()

                            mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(get_data['class'],period,day))
                            mydb.commit()        
                        
                    elif(get_id_data.type=="PROJECT"):

                        get_faculty_split=(get_id_data.faculty).split(",")

                        day_period_split=j.split('--')
                        day=day_period_split[0]
                        period=day_period_split[1]

                        for fac in get_faculty_split:

                            mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(fac,period,day))
                            mydb.commit()

                            work=Teacheradd.query.filter_by(name=fac).first()

                            work.work_load-=1
                            db.session.commit()

                            mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(get_data['class'],period,day))
                            mydb.commit() 

                    else:#electives

                        get_faculty=get_id_data.faculty

                        work=Teacheradd.query.filter_by(name=get_faculty).first()

                        day_period_split=j.split('--')
                        day=day_period_split[0]
                        period=day_period_split[1]

                        mycursor.execute("SELECT `{1}` FROM `{0}` WHERE DAY='{2}';".format(get_faculty,period,day))


                        #split selected by '/'

                        print("select completed")
                        val=mycursor.fetchall()

                        val=(val[0][0]).split('/')

                        print(val)

                        if(len(val)==2):

                            mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(get_faculty,period,day))
                            mydb.commit()

                            #mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(get_data['class'],period,day))
                            mydb.commit()

                            work.work_load-=1

                            db.session.commit()

                        else:
                            #just update faculty cell by removing current class and don't decrease the workload
                            #val.remove(get_data['data'])

                            val.remove(get_data['class'])

                            delimeter="/"

                            print("fac_cell",val)

                            st=('/').join(val)

                            print(st)

                            mycursor.execute("UPDATE `{0}` SET `{1}`='{2}' WHERE `Day`='{3}';".format(get_faculty,period,st,day))
                            mydb.commit()

                            print("elective string:",st)

                        #res=session.query(classconfig_table).filter(classconfig_table.classname==get_data['class'],classconfig_table.type==get_id_data.type)

                        res=classconfig_table.query.filter_by(classname=get_data['class'],type=get_id_data.type)

                        c=0
                        for k in res:
                            print("check single elective delete day period --",k.dayperiod.split(','))

                            print('subject',k.subject)

                            if(j in k.dayperiod.split(',')):
                                c=c+1
                        print("count",c)

                        if(c==1):

                            mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(get_data['class'],period,day))
                            mydb.commit()

                        
                
                if(flag==0):
                    get_id_data.allocated=0
                    db.session.commit()

                    get_id_data.dayperiod='@'
                    db.session.commit() 

                    spl=(get_id_data.split).split('/')
                    get_id_data.split=spl[0]
                    db.session.commit()
                    

        else:# only consider  day/from/to dayperiod=(get_id_data.dayperiod).split(",")

            id_split=get_data['ids'].split(',')

            if(len(id_split)==1):
                print("single id only consider  day/from/to")

                if(int(get_data['from'])<=int(get_data['to'])):

                    get_id_data=classconfig_table.query.get(int(id_split[0]))

                    dayperiod=(get_id_data.dayperiod).split(",")

                    get_day_period=get_id_data.dayperiod
                    
                    if(get_day_period=="@"):
                        return jsonify( {'msg': "Periods not been assigned at these periods and day"})
                    else:

                        spl=(get_id_data.split).split('/')

                        if(str((int(get_data['to'])+1)-int(get_data['from'])) not in spl[1] ):

                            return jsonify({'msg':"Given scale does not fitted to any split period"})

                        get_day_period=(get_id_data.dayperiod).split(',')

                        #check both input is present in id's day period

                        st=[]#it contain selected scale of day--period
                        for i in range(int(get_data['from']),int(get_data['to'])+1):
                            st.append(get_data['day']+"--"+str(i))

                        print(st)

                        print(set(st).intersection(set(get_day_period)))

                        if(set(st).intersection(set(get_day_period))==set(st)):

                            print("day_period match")
                            store=get_day_period
                            # j is day--period
                            for j in st:

                                day_period_split=j.split('--')
                                day=day_period_split[0]
                                period=day_period_split[1] 

                                if(get_id_data.type=="THEORY"):

                                    get_faculty=get_id_data.faculty
                                    work=Teacheradd.query.filter_by(name=get_faculty).first()
                    
                                    mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(get_faculty,period,day))
                                    mydb.commit()   

                                    mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(get_data['class'],period,day))
                                    mydb.commit()   

                                    work.work_load-=1

                                    db.session.commit() 

                                    get_id_data.allocated-=1
                                    db.session.commit()
                                    

                                elif(get_id_data.type=="LAB"):

                                    get_faculty_split=(get_id_data.faculty).split(",")

                                    get_lab_split=(get_id_data.lab).split(",")


                                    for fac in get_faculty_split:

                                        mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(fac,period,day))
                                        mydb.commit()

                                        work=Teacheradd.query.filter_by(name=fac).first()

                                        work.work_load-=1
                                        db.session.commit()

                                    for lab in get_lab_split:

                                        mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(lab,period,day))
                                        mydb.commit()

                                        mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(get_data['class'],period,day))
                                        mydb.commit()        
                                    get_id_data.allocated-=1
                                    db.session.commit()  


                                elif(get_id_data.type=="PROJECT"):

                                    get_faculty_split=(get_id_data.faculty).split(",")


                                    for fac in get_faculty_split:

                                        mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(fac,period,day))
                                        mydb.commit()

                                        work=Teacheradd.query.filter_by(name=fac).first()

                                        work.work_load-=1
                                        db.session.commit()

                                        mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(get_data['class'],period,day))
                                        mydb.commit() 

                                    get_id_data.allocated-=1
                                    db.session.commit()

                                else:#electives

                                    print("elective delete by scale")
                                    get_faculty=get_id_data.faculty

                                    work=Teacheradd.query.filter_by(name=get_faculty).first()

                                    mycursor.execute("SELECT `{1}` FROM `{0}` WHERE DAY='{2}';".format(get_faculty,period,day))

                                    #split selected by '/'

                                    print("select completed")
                                    val=mycursor.fetchall()

                                    val=(val[0][0]).split('/')

                                    print(val)

                                    if(len(val)==2):

                                        mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(get_faculty,period,day))
                                        mydb.commit()

                                        work.work_load-=1

                                        db.session.commit()                                    

                                        """ele=classconfig_table.query.filter_by(type=get_id_data.type,classname=get_data['class'])

                                        ele_=[]
                                        for p in ele:
                                            ele_.append((p.dayperiod).split(','))

                                        ele_=sum(ele_, [])
                                        print(j,ele_)

                                        if(ele_.count(j)==1):                                    

                                            mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(get_data['class'],period,day))
                                            mydb.commit()"""

                                    else:
                                        #just update faculty cell by removing current class and don't decrease the workload
                                        #val.remove(get_data['data'])

                                        if(get_data['class'] in val):

                                            val.remove(get_data['class'])

                                        delimeter="/"

                                        print("fac_cell",val)

                                        string=('/').join(val)

                                        print(string)

                                        mycursor.execute("UPDATE `{0}` SET `{1}`='{2}' WHERE `Day`='{3}';".format(get_faculty,period,string,day))
                                        mydb.commit()

                                        #ele=session.query(classconfig_table).filter(classconfig_table.type==get_id_data.type,classconfig_table.classname==get_data['class'])


                                    ele=classconfig_table.query.filter_by(type=get_id_data.type,classname=get_data['class']).all()


                                    """#reg=str('%'+str(day)+'%')
                                    reg=str(day)+"%"
                                    print("day",day)
                                    mycursor.execute("SELECT dayperiod classconfig_table FROM WHERE type='{0}' AND classname='{1}'; ".format(get_id_data.type,get_data['class']))

                                    ele=mycursor.fetctall()"""
                                    count_delete=0
                                    for li in ele:

                                        print(li.dayperiod)

                                        if(j in li.dayperiod.split(',')):
                                            count_delete+=1

                                    """ele_=[]
                                    for p in ele:
                                        ele_.append((p.dayperiod).split(','))

                                    ele_=sum(ele_, [])
                                    print(j,ele_)"""

                                    print("count",len(ele))

                                    if(count_delete==1): 

                                        print("came to if ")
                                        mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(get_data['class'],period,day))
                                        mydb.commit()

                                        print("elective string:",st) 

                                        """res=classconfig_table.query.filter_by(classname=get_data['class'],type=get_id_data.type)

                                        c=0
                                        for k in res:
                                            print("check single elective delete day period --",k.dayperiod.split(','))

                                            print('subject',k.subject)

                                            if(j in k.dayperiod.split(',')):
                                                c=c+1
                                        print("count",c)
                                        
                                        if(c==1):

                                            mycursor.execute("UPDATE `{0}` SET `{1}`='--' WHERE `Day`='{2}';".format(get_data['class'],period,day))
                                            mydb.commit()"""

                                    get_id_data.allocated-=1
                                    db.session.commit()


                                    store.remove(j)

                                    if(store):
                                        get_id_data.dayperiod=(',').join(store)
                                        db.session.commit()
                                    else:
                                        get_id_data.dayperiod="@"
                                        db.session.commit() 


                                if(get_id_data.type=="THEORY" or get_id_data.type=="LAB" or  get_id_data.type=="PROJECT" ):
                                    store.remove(j)

                                    if(store):
                                        get_id_data.dayperiod=(',').join(store)
                                        db.session.commit()
                                    else:
                                        get_id_data.dayperiod="@"
                                        db.session.commit()


                            #update id split after for loop
                            if(get_id_data.allocated==0):
                                spl=(get_id_data.split).split('/')
                                get_id_data.split=spl[0]
                                db.session.commit()
                            else:
                                scale=str((int(get_data['to'])+1)-int(get_data['from']))
                                spl=(get_id_data.split)#list formed 

                                print(spl,spl[0],spl[1],scale)

                                
                                index=spl.find('/')

                                first=spl[:index]

                                second=spl[index:]

                                second_split=second.split(',')

                                second_split.remove(scale)

                                second=(',').join(second_split)

                                
                                get_id_data.split=first+second
                                db.session.commit()


                        else:

                            return jsonify( {'msg': "periods/Day scale is not present in subject's column"})                       


                else:
                    
                    return jsonify( {'day': "'from period' should less than equal to 'To day"})
                
            else:
                print("------>only one is allowed")


                return jsonify( {'msg': "only one checkbox is allowed"})

        get_period=day_period.query.get(1)

        if(get_period):
            mycursor.execute("SELECT * FROM `{0}`;".format(get_data['class']))
            row=mycursor.fetchall()
            all_data=classconfig_table.query.all()

            return jsonify( {'data': render_template('classtable.html',all_data=row,periods=get_period.periods),'describe':render_template('classtable_describe.html',classname=get_data['class'],all_data=all_data)})
    else:
        return redirect(url_for('dashboard'))

@app.route('/delete_all_classes', methods = ['GET', 'POST'])
def delete_all_classes():
    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table')) 
        #get_data=request.get_json()

        #print(get_data)


        #all_faculty=classconfig_table.query.with_entities(classconfig_table.faculty).distinct()
        mycursor = mydb.cursor(buffered=True)
        l1=[]

        all_id=classconfig_table.query.with_entities(classconfig_table.id).distinct()

        all_faculty=Teacheradd.query.all()

        for i in all_id:

            ids=classconfig_table.query.get(i[0])
            ids.allocated=0
            db.session.commit()
            ids.dayperiod="@"
            db.session.commit()

            first=(ids.split).split('/')

            ids.split=first[0]
            db.session.commit()


            if(ids.type=="LAB" or ids.type=="PROJECT"):
                ids.faculty="NA"
                db.session.commit()  


        for i in all_faculty:
            print(i.name)
            i.work_load=0
            db.session.commit()    


        #all_lab=classconfig_table.query.with_entities(classconfig_table.lab).distinct()
        all_lab=Labs.query.with_entities(Labs.name).distinct()

        all_class=classconfig_table.query.with_entities(classconfig_table.classname).distinct()

        for i in all_faculty:

            l1.append((i.name).split(","))

        for i in all_lab:
            l1.append(i[0].split(","))

        for i in all_class:
            l1.append(i[0].split(","))

        flat_set = set(sum(l1, []))

        print("flat_set",flat_set)
        get_period=day_period.query.filter_by(id=1).first()
        print(get_period.periods)

        for i in flat_set:
            flag=show_all_tables(i.lower())
            print(flag)
            if(flag==True):

                for j in range(1,get_period.periods+1):

                    mycursor.execute("UPDATE `{0}` SET `{1}`='--' ".format(i,j))
                    mydb.commit()

                #deleteF(i)
                #create_table(i,7,6)
            else:
                print(i,"table is not there")
        
            
        return jsonify( {'data':"success"})
    else:
	        
        """mycursor.execute("DELETE FROM `{0}`".format(get_data['class']))

        deleteF(get_data['clss'])
        create_table(get_data['class'])
        mydb.commit()


        mycursor.execute("SELECT * FROM `{0}`;".format(get_data['class']))
        row=mycursor.fetchall()
        print(row)
        return jsonify( {'data': render_template('classtable.html',all_data=row )})"""
        return redirect(url_for('dashboard'))


@app.route('/day_wise')
def day_wise():
    if 'loggedin' in session:
        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table')) 

        mycursor = mydb.cursor(buffered=True)
        day_list=['MON','TUE','WED','THU','FRI','SAT','SUN']

        get_day=day_period.query.get(1)
        day_list=day_list[:get_day.day]

        all_course=[]
        all_dept=department.query.with_entities(department.name).distinct()

        all_course=Courseadd.query.all()
        
        final=[]
        for co in all_course:
            l1=[]
            l1.append(co.id)
            l1.append(co.name)
            l1.append(co.semisters)
            l1.append(co.dept)
            final.append(l1)
            
        return render_template("day_wise.html",day_list=day_list,access=session['type'],all_course=final,all_dept=all_dept)
    else:

	    return redirect(url_for('dashboard'))

@app.route('/period_wise')
def period_wise():
    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table')) 

        day_list=['MON','TUE','WED','THU','FRI','SAT','SUN']
        
        mycursor = mydb.cursor(buffered=True)
        get_period=day_period.query.get(1)
        day_list=day_list[:get_period.day]

        all_course=Courseadd.query.all()

        final=[]
        for co in all_course:
            l1=[]
            l1.append(co.id)
            l1.append(co.name)
            l1.append(co.semisters)
            l1.append(co.dept)
            final.append(l1)

        all_dept=department.query.with_entities(department.name).distinct()

        return render_template("period_wise.html",periods=get_period.periods,day_list=day_list,access=session['type'],all_course=final,all_dept=all_dept)
    else:

	    return redirect(url_for('dashboard'))

@app.route('/free_period_list')
def free_period_list():
    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table')) 

        mycursor = mydb.cursor(buffered=True)
        day_list=['MON','TUE','WED','THU','FRI','SAT','SUN']   
        get_day=day_period.query.get(1)
        day_list=day_list[:get_day.day]

        all_dept=department.query.with_entities(department.name).distinct()
        all_course=Courseadd.query.all()
        final=[]
        for co in all_course:
            l1=[]
            l1.append(co.id)
            l1.append(co.name)
            l1.append(co.semisters)
            l1.append(co.dept)
            final.append(l1)


        return render_template("free_period_list.html",day_list=day_list,access=session['type'],all_course=final,all_dept=all_dept)
    else:

	    return redirect(url_for('dashboard'))

@app.route('/view_fac_lab_tt')
def view_fac_lab_tt():

    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table')) 

        mycursor = mydb.cursor(buffered=True)

        all_dept=Teacheradd.query.with_entities(Teacheradd.branch).distinct()
        all_teacher=Teacheradd.query.all()
        all_lab=Labs.query.with_entities(Labs.name).distinct()
        l1=[]
        for i in all_lab:
            l1.append(i[0].split(","))

        all_lab=set(sum(l1,[]))

        if("NA" in all_lab):
            all_lab.remove("NA")
        
            
        list_total_faculty=[]
        for i in all_teacher:
                list_faculty=[]
                list_faculty.append(i.name)
                list_faculty.append(i.branch)

                list_total_faculty.append(list_faculty)
    
        return render_template("view_fac_lab_tt.html",all_teacher=list_total_faculty,all_dept=all_dept,all_lab=all_lab,access=session['type'])

    else: 
        return redirect(url_for('dashboard'))

@app.route('/get_report', methods = ['GET', 'POST'])
def get_report():

    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))


        mycursor = mydb.cursor(buffered=True)

        get_data=request.get_json()

        print(get_data)

        all_id=classconfig_table.query.with_entities(classconfig_table.id).distinct()

        l1=[]
        """for i in all_class:
            l1.append(i.classname)"""

        table=[]
        print(l1)

        get_day_period=day_period.query.get(1)

        print(get_day_period.day,get_day_period.periods)

        #get_course=Courseadd.query.filter_by(dept=get_data['dept']).all()

        if(get_data['status']=="day_wise"):
            if(get_data['day'] and not get_data['dept'] and not get_data['sem'] and not get_data['course']):
                #display all
                print("only day")

                all_class=classconfig_table.query.with_entities(classconfig_table.classname).distinct()
                
                for i in all_class:
                    l1.append(i.classname)

                all_lab=Labs.query.with_entities(Labs.name).distinct()
                for i in all_lab:
                    l1.append(i.name)

            elif(get_data['day'] and get_data['dept'] and not get_data['course'] and not get_data['sem'] ):
                #display all classes of selected dept

                get_course=Courseadd.query.filter_by(dept=get_data['dept']).all()

                all_class=[]
                for course in get_course:
                #select class and their respective class
        
                    reg=course.name+'%'
            
                    mycursor.execute("SHOW TABLES WHERE Tables_in_{0} LIKE '{1}'; ".format("jntuk1",reg))

                    get_classes=mycursor.fetchall()

                    for cls in get_classes:
                        all_class.append(cls[0])

                    print(all_class)

                all_lab=Labs.query.filter_by(branch=get_data['dept']).all()  

                for i in all_class:
                    l1.append(i)

                for i in all_lab:
                    l1.append(i.name)

            elif(get_data['day'] and get_data['dept'] and  get_data['course'] and not get_data['sem'] ):
                #display all classes of selected dept and select course
                all_class=[]

                reg=get_data['course']+'%'
            
                mycursor.execute("SHOW TABLES WHERE Tables_in_{0} LIKE '{1}'; ".format("jntuk1",reg))

                get_classes=mycursor.fetchall()

                print(get_classes)
                for cls in get_classes:
                    l1.append(cls[0])
                all_lab=Labs.query.filter_by(branch=get_data['dept']).all()
                for i in all_lab:
                    l1.append(i.name)

            elif(get_data['day'] and get_data['dept'] or  get_data['course'] and  get_data['sem'] ):
                print("else came")
                #select odd or even classes
                all_class=[]

                get_course=Courseadd.query.filter_by(dept=get_data['dept']).all()
                for course in get_course:
                #select class and their respective class

                    print(course.name)
                    reg=course.name+'%'
            
                    mycursor.execute("SHOW TABLES WHERE Tables_in_{0} LIKE '{1}'; ".format("jntuk1",reg))

                    get_classes=mycursor.fetchall() 

                    for cls in get_classes:
                        all_class.append(cls[0])

                print(get_data['sem'])

                if(get_data['sem']=='Odd'):

                    for cls in all_class:
                        print("cls:",cls,cls[-1])
                        if(int(cls[-1])%2==1):
                            l1.append(cls)

                else:
                    for cls in all_class:
                        print("cls:",cls[-1])
                        if(int(cls[-1])%2==0):
                            l1.append(cls)

                all_lab=Labs.query.filter_by(branch=get_data['dept']).all()
                for i in all_lab:
                    l1.append(i.name)

            else:
                pass
            
            print("final table list:",l1)
            for i in l1:
                flag=show_all_tables(i.lower())
                row=[]
                row.append(i)
                if(flag==True):
                    for j in range(1,get_day_period.periods+1):

                        mycursor.execute("SELECT `{0}` FROM `{1}` WHERE DAY='{2}';".format(j,i,get_data['day']))

                        val=mycursor.fetchall()
                        row.append(val[0][0])
                        #print(val[0][0])
                    table.append(row)
            

               
            return jsonify({'describe':render_template('day_wise_table.html',all_data=table,periods=get_day_period.periods,access=session['type'])})

        elif(get_data["status"]=="period_wise"):

            day_list=['MON','TUE','WED','THU','FRI','SAT','SUN']

            get_day=day_period.query.get(1)

            day_list=day_list[:get_day.day]

            length=len(day_list)

            if(get_data['period'] and not get_data['dept'] and not get_data['sem'] and not get_data['course']):
                #display all
                print("only day")

                all_class=classconfig_table.query.with_entities(classconfig_table.classname).distinct()
                
                for i in all_class:
                    l1.append(i.classname)

                all_lab=Labs.query.with_entities(Labs.name).distinct()
                for i in all_lab:
                    l1.append(i.name)

            elif(get_data['period'] and get_data['dept'] and not get_data['course'] and not get_data['sem'] ):
                #display all classes of selected dept

                get_course=Courseadd.query.filter_by(dept=get_data['dept']).all()

                all_class=[]
                for course in get_course:
                #select class and their respective class
        
                    reg=course.name+'%'
            
                    mycursor.execute("SHOW TABLES WHERE Tables_in_{0} LIKE '{1}'; ".format("jntuk1",reg))

                    get_classes=mycursor.fetchall()

                    for cls in get_classes:
                        all_class.append(cls[0])

                    print(all_class)

                all_lab=Labs.query.filter_by(branch=get_data['dept']).all()  

                for i in all_class:
                    l1.append(i)

                for i in all_lab:
                    l1.append(i.name)

            elif(get_data['period'] and get_data['dept'] and  get_data['course'] and not get_data['sem'] ):
                #display all classes of selected dept and select course
                all_class=[]

                reg=get_data['course']+'%'
            
                mycursor.execute("SHOW TABLES WHERE Tables_in_{0} LIKE '{1}'; ".format("jntuk1",reg))

                get_classes=mycursor.fetchall()

                print(get_classes)
                for cls in get_classes:
                    l1.append(cls[0])
                all_lab=Labs.query.filter_by(branch=get_data['dept']).all()
                for i in all_lab:
                    l1.append(i.name)

            elif(get_data['period'] and get_data['dept'] or  get_data['course'] and  get_data['sem'] ):
                print("else came")
                #select odd or even classes
                all_class=[]

                get_course=Courseadd.query.filter_by(dept=get_data['dept']).all()
                for course in get_course:
                #select class and their respective class

                    print(course.name)
                    reg=course.name+'%'
            
                    mycursor.execute("SHOW TABLES WHERE Tables_in_{0} LIKE '{1}'; ".format("jntuk1",reg))

                    get_classes=mycursor.fetchall() 

                    for cls in get_classes:
                        all_class.append(cls[0])

                print(get_data['sem'])

                if(get_data['sem']=='Odd'):

                    for cls in all_class:
                        print("cls:",cls,cls[-1])
                        if(int(cls[-1])%2==1):
                            l1.append(cls)

                else:
                    for cls in all_class:
                        print("cls:",cls[-1])
                        if(int(cls[-1])%2==0):
                            l1.append(cls)

                all_lab=Labs.query.filter_by(branch=get_data['dept']).all()
                for i in all_lab:
                    l1.append(i.name)

            else:
                pass            


            for i in l1:
                flag=show_all_tables(i.lower())
                row=[]
                row.append(i)
                if(flag==True):
                    for j in day_list:
                        mycursor.execute("SELECT `{0}` FROM `{1}` WHERE Day='{2}';".format(get_data['period'],i,j))

                        val=mycursor.fetchall()
                        row.append(val[0][0])

                    table.append(row)

            return jsonify({'describe':render_template('period_wise_table.html',all_data=table,day_list=day_list,length=length,access=session['type'])})
        elif(get_data['status']=="free_periods"):

            print("faculty free periods entered")
            if( not get_data['dept']):

                fac_list=Teacheradd.query.order_by(Teacheradd.work_load).all()


                for i in fac_list:
                    flag=show_all_tables((i.name).lower())
                    row=[]
                    row.append(i.name+'('+str(i.work_load)+')')
                    if(flag==True):
                        for j in range(1,get_day_period.periods+1):
                            mycursor.execute("SELECT `{0}` FROM `{1}` WHERE Day='{2}';".format(str(j),i.name,get_data['day']))

                            val=mycursor.fetchall()
                            row.append(val[0][0])

                        table.append(row)

                return jsonify({'describe':render_template('day_wise_table.html',all_data=table,periods=get_day_period.periods,access=session['type'])})
            else:
                all_fac=Teacheradd.query.order_by(Teacheradd.work_load).all()

                dept_fac=[]
                for fac in all_fac:
                    if(fac.branch==get_data['dept']):

                        dept_fac.append([fac.name,fac.work_load])

                for i in dept_fac:
                    flag=show_all_tables((i[0]).lower())
                    row=[]
                    
                    row.append(i[0]+'('+str(i[1])+')')
                    if(flag==True):
                        for j in range(1,get_day_period.periods+1):
                            mycursor.execute("SELECT `{0}` FROM `{1}` WHERE Day='{2}';".format(str(j),i[0],get_data['day']))

                            val=mycursor.fetchall()
                            row.append(val[0][0])

                        table.append(row)

                return jsonify({'describe':render_template('day_wise_table.html',all_data=table,periods=get_day_period.periods,access=session['type'])})

        else:
            pass
    else:
       return redirect(url_for('dashboard')) 


@app.route('/classtable_get', methods = ['GET', 'POST'])
def classtable_get():
    if 'loggedin' in session:
        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))

        get_data=request.get_json()
        print(get_data)

        mycursor = mydb.cursor(buffered=True)

        mycursor.execute("SELECT * FROM `{0}`;".format(get_data['table']))
        row=mycursor.fetchall()

        all_data=classconfig_table.query.all()

        get_periods=day_period.query.get(1)
        
        if(get_data["status"]=="class"):

            return jsonify( {'data': render_template('classtable.html',all_data=row,periods=get_periods.periods),'describe':render_template('classtable_describe.html',classname=get_data['table'],all_data=all_data)})

        elif(get_data["status"]=="lab"):
            return jsonify( {'data': render_template('classtable.html',all_data=row,periods=get_periods.periods ),'describe':render_template('lab_table_describe.html',lab=get_data['table'],all_data=all_data)})
        
        elif(get_data["status"]=="faculty"):
            return jsonify( {'data': render_template('classtable.html',all_data=row,periods=get_periods.periods ),'describe':render_template('faculty_table_describe.html',faculty=get_data['table'],all_data=all_data)})
    else:
        return redirect(url_for('dashboard'))

@app.route('/classtable_entry', methods = ['GET', 'POST'])
def classtable_entry():

    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))


        print("classtable_entry configuration")

        mycursor = mydb.cursor(buffered=True)   
        get_data=request.get_json()

        get_period=day_period.query.get(1)
        periods=get_period.periods
        print(get_data)

        data=classconfig_table.query.get(get_data['id'])

        print(data.allocated, data.total_periods)

        if(str((int(get_data['to'])+1)-int(get_data['from'])) not in (data.split).split(',')):

            return jsonify({'max':"Given scale does not fitted to any split period"})

        if(get_data['type']=="THEORY" or get_data['type'][:9]=="ELECTIVE/"):

            print("class theory  configuration")

            #str1="SELECT `{0}` FROM `{1}` WHERE `{2}`=`{0}` ".format(str(get_data['period']),str(get_data['faculty']),str(get_data['day']))
            #mycursor.execute("SELECT `2` FROM dhanalakshmi WHERE Day='M0N';")
            #rows=mycursor.fetchall()

            #print(rows)
            mycursor = mydb.cursor(buffered=True)

            #print(type(get_data['from']))

            to=int(get_data['to'])
            frm=int(get_data['from'])
            counter=frm-1

            print(frm,to,counter)

            print(str(to)==get_data['to'])
            print(str(frm)==get_data['from'])
            
            print("faculty name",get_data['faculty'])



            if(frm<to+1):

                if(data.allocated+((int(get_data['to'])+1)-int(get_data['from']))<=data.total_periods):

                    #given scale in split or not



                    if(get_data['type'][:9]== "ELECTIVE/"):


                        #get

                        get_list_=classconfig_table.query.filter_by(subject=get_data['subject'],type=get_data['type']).all()

                        for p in get_list_:
                            print("list class",p.classname)


                        #get_list_class=session.query(classconfig_table).filter(classconfig_table.subject==get_data['subject'],classconfig_table.type==get_data['type'])

                        #print(get_list_class)

                        class_list=[]
                        for p in get_list_:
                            print(p.classname)
                            class_list.append(p.classname)

                        print("class_list",class_list)

                        for i in range(frm,to+1):

                            #select faculty
                            mycursor.execute("SELECT `{0}` FROM `{1}` WHERE Day='{2}';".format(str(i),get_data['faculty'],get_data['day']))

                            fac=mycursor.fetchall()
                            print(fac)

                            free=fac[0][0]

                            mycursor.execute("SELECT `{0}` FROM `{1}` WHERE Day='{2}';".format(str(i),get_data['class'],get_data['day']))

                            cls=mycursor.fetchall()
                            print(cls)     

                            print(len(fac[0][0]),fac[0][0],len(cls[0][0]),cls[0][0])

                            print("(fac[0][0] =='--' or get_data['faculty'] in fac_list )")


                            if(i==frm):
                                old=cls[0][0] 

                            if((fac[0][0] =='--' or ( get_data['class'] in class_list and get_data['class'] not in fac[0][0].split("/")) ) and (cls[0][0]=='--' or cls[0][0]==get_data['type'])):


                                if(cls[0][0]==old):
                                    counter=counter+1    

                            print("counter",counter)               

                    else:#theory configuration              

                        for i in range(frm,to+1):


                            mycursor.execute("SELECT `{0}` FROM `{1}` WHERE Day='{2}';".format(str(i),get_data['faculty'],get_data['day']))

                            fac=mycursor.fetchall()
                            print(fac)

                            

                            mycursor.execute("SELECT `{0}` FROM `{1}` WHERE Day='{2}';".format(str(i),get_data['class'],get_data['day']))

                            cls=mycursor.fetchall()
                            print(cls)

                            print(len(fac[0][0]),fac[0][0],len(cls[0][0]),cls[0][0])

                            if(fac[0][0] =='--' and cls[0][0]=='--'):  
                                counter=counter+1
                    

                    if(counter==int(get_data['to'])):#check it once

                        
                        
                        if( get_data['type'][:9]== "ELECTIVE/" ): 

                            for i in range(frm,to+1):

                                #update faculty

                                if(free=="--"):

                                    mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE `Day`='{2}';".format(str(i),get_data['faculty'],get_data['day'],get_data['class']+'/'+get_data['subject']))
                                    mydb.commit()

                                    #update class
                                    mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE Day='{2}';".format(str(i),get_data['class'],get_data['day'],get_data['type']))
                                    mydb.commit()

                                    faculty=Teacheradd.query.filter_by(name=get_data['faculty']).first()

                                    faculty.work_load=faculty.work_load+1      
                                    db.session.commit()                          

                                else:
                                    mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE `Day`='{2}';".format(str(i),get_data['faculty'],get_data['day'],get_data['class']+"/"+free))
                                    mydb.commit()

                                    #update class
                                    mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE Day='{2}';".format(str(i),get_data['class'],get_data['day'],get_data['type']))
                                    mydb.commit()

                            
                                data.allocated=data.allocated+1
                                db.session.commit()

                                if(data.dayperiod=="@"):
                                    data.dayperiod=get_data['day']+"--"+str(i)
                                    db.session.commit()
                                else:
                                    data.dayperiod= data.dayperiod +","+get_data['day']+"--"+str(i)
                                    db.session.commit()


                        else:#theory   subjects update 
                            for i in range(frm,to+1):   
                                mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE `Day`='{2}';".format(str(i),get_data['faculty'],get_data['day'],get_data['class']+"/"+get_data['subject']))
                                mydb.commit()

                                mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE Day='{2}';".format(str(i),get_data['class'],get_data['day'],get_data['subject']))
                                mydb.commit()

                                data.allocated=data.allocated+1
                                db.session.commit()

                                faculty=Teacheradd.query.filter_by(name=get_data['faculty']).first() 

                                faculty.work_load=faculty.work_load+1
                                db.session.commit()

                                if(data.dayperiod=="@"):
                                    data.dayperiod=get_data['day']+"--"+str(i)
                                    db.session.commit()
                                else:
                                    data.dayperiod= data.dayperiod +","+get_data['day']+"--"+str(i)
                                    db.session.commit()
                    

                    else:
                        return jsonify({'collide':"slot is already assigned in this fac/class"})
                else:

                    return jsonify({'max':"maximum allotment of periods satisfied"})
            else:
                return jsonify({'day':" 'from day' should less than equal to 'to day' "})

        elif(get_data['type']=='PROJECT'):
            print("project configuration")

    

            mycursor = mydb.cursor(buffered=True)

            #print(type(get_data['from']))

            to=int(get_data['to'])
            frm=int(get_data['from'])

            print(frm,to)

            print(str(to)==get_data['to'])
            print(str(frm)==get_data['from'])
            
                
            data=classconfig_table.query.get(get_data['id'])

            fac_split=get_data['faculty'].split(",")

            flag=0

            for i in fac_split:

                if(fac_split.count(i)!=1):
                    flag=1
            

            if((to+1)-frm + data.allocated <= data.total_periods ):

                if(data.faculty=='NA'):

                    if(flag==0):

                        if(len(fac_split)==data.f_count):
                            
                            if(frm<to+1):
                                #checking availabilty of class slots
                                for i in range(frm,to+1):
                                    mycursor.execute("SELECT `{0}` FROM `{1}` WHERE Day='{2}';".format(str(i),get_data['class'],get_data['day']))

                                    cls=mycursor.fetchall()
                                    print(cls)

                                    print(len(cls[0][0]),cls[0][0])

                                    if(not cls[0][0]=='--'):  

                                        return jsonify({'class_occupied':"class is already occupied in this slot range"})
                                #endfor

                                for i in fac_split:
                                    #select for "--" in fac list

                                    for j in range(frm,to+1):

                                        mycursor.execute("SELECT `{0}` FROM `{1}` WHERE Day='{2}';".format(str(j),i,get_data['day']))

                                        cls=mycursor.fetchall()

                                        print(cls)

                                        print(len(cls[0][0]),cls[0][0])

                                        if(not cls[0][0]=='--'):  

                                            return jsonify({'fac_occupied':"faculty is already occupied in this slot range"})
                                        #endfor inner                             
                                #endfor

                                #update class
                                for i in range(frm,to+1):
                                    mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE Day='{2}';".format(str(i),get_data['class'],get_data['day'],get_data['subject']))
                                    mydb.commit()

                                    data.allocated=data.allocated+1
                                    db.session.commit()  

                                    if(data.dayperiod=="@"):
                                        data.dayperiod=get_data['day']+"--"+str(i)
                                        db.session.commit()
                                    else:
                                        data.dayperiod= data.dayperiod +","+get_data['day']+"--"+str(i)
                                        db.session.commit() 


                                #update faculty  list

                                for j in fac_split:
                                    faculty=Teacheradd.query.filter_by(name=j).first()

                                    
                                    for i in range(frm,to+1):
                                        mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE `Day`='{2}';".format(str(i),j,get_data['day'],get_data['class']+"/"+get_data['subject']))
                                        mydb.commit()     

                                        faculty.work_load=faculty.work_load+1
                                        db.session.commit()
                                    

                                #endfor 
                                
                                data.faculty=get_data['faculty']
                                db.session.commit()

                            else:
                                return jsonify({'day_error':"from_day should less than to_day"})

                        else:
                            return jsonify({'min_error':"enough count of faculty is not satisfied"})
                    else:
                        return jsonify({'duplicate_faclty':"faculty should not be repeated"})
                else:
                    
                    if(flag==0):

                        data_fac_split=data.faculty.split(",")

                        if(set(data_fac_split)==set(fac_split)):

                            if(frm<to+1):
                                #checking availabilty of class slots
                                for i in range(frm,to+1):
                                    mycursor.execute("SELECT `{0}` FROM `{1}` WHERE Day='{2}';".format(str(i),get_data['class'],get_data['day']))

                                    cls=mycursor.fetchall()
                                    print(cls)

                                    print(len(cls[0][0]),cls[0][0])

                                    if(not cls[0][0]=="--"):  

                                        return jsonify({'class_occupied':"class is already occupied in this slot range"})
                                #endfor

                                for i in fac_split:
                                    #select for "--" in fac list

                                    for j in range(frm,to+1):

                                        mycursor.execute("SELECT `{0}` FROM `{1}` WHERE Day='{2}';".format(str(j),i,get_data['day']))

                                        cls=mycursor.fetchall()

                                        print(cls)

                                        print(len(cls[0][0]),cls[0][0])

                                        if(not cls[0][0]=='--'):  

                                            return jsonify({'fac_occupied':"faculty is already occupied in this slot range"})
                                    #endfor inner                             
                                #endfor

                                #update class
                                for i in range(frm,to+1):

                                    mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE Day='{2}';".format(str(i),get_data['class'],get_data['day'],get_data['subject']))
                                    mydb.commit()

                                    data.allocated=data.allocated+1
                                    db.session.commit()  

                                    if(data.dayperiod=="@"):
                                        data.dayperiod=get_data['day']+"--"+str(i)
                                        db.session.commit()
                                    else:
                                        data.dayperiod= data.dayperiod +","+get_data['day']+"--"+str(i)
                                        db.session.commit() 


                                #update faculty  list

                                for j in fac_split:
                                    faculty=Teacheradd.query.filter_by(name=j).first()

                                    for i in range(frm,to+1):
                                        mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE `Day`='{2}';".format(str(i),j,get_data['day'],get_data['class']+'/'+get_data['subject']))
                                        mydb.commit()     

                                        faculty.work_load=faculty.work_load+1
                                        db.session.commit() 

                                #endfor 
                            else:
                                return jsonify({'day_error':"from_day should less than to_day"})
                                
                        else:
                            return jsonify({'duplicate_faclty':"faculty should  be UNIQUE"})
                    else:
                        return jsonify({'duplicate_faclty':"faculty should not be repeated"})
            else:
                return jsonify({'max':"maximum allotment of periods satisfied"})

        else:#lab_configure

            print("lab configuration ALREADY FACULTY PRESENT ")
            mycursor = mydb.cursor(buffered=True)

            #print(type(get_data['from']))

            to=int(get_data['to'])
            frm=int(get_data['from'])

            print(frm,to)

            print(str(to)==get_data['to'])
            print(str(frm)==get_data['from'])
            
                
            data=classconfig_table.query.get(get_data['id'])

            fac_split=get_data['faculty'].split(",")

            print("fac_split",fac_split)

            flag=0

            for i in fac_split:

                if(fac_split.count(i)!=1 ):#check duplicate faculty exist or not
                    flag=1
            
            res=(to+1)-frm + data.allocated
            print("res:",res,'allocated',data.allocated)

            if((((to+1)-frm) + data.allocated) <= data.total_periods ):

                if(data.faculty=='NA'):
                    print(data.faculty)
                    if(flag==0):
                        print(flag)
                        if(len(fac_split)==data.f_count):
                            print(len(fac_split)==data.f_count)
                            if(frm<to+1):
                                print(frm<to+1)
                                #checking availabilty of class slots

                                for i in range(frm,to+1):
                                    mycursor.execute("SELECT `{0}` FROM `{1}` WHERE Day='{2}';".format(str(i),get_data['class'],get_data['day']))

                                    cls=mycursor.fetchall()
                                    print(cls)

                                    print(len(cls[0][0]),cls[0][0])

                                    if(not (cls[0][0]=='--')):  

                                        return jsonify({'class_occupied':"class {0} already occupied in this slot range".format(get_data['class'])})
                                #endfor

                                #checking availability in lab slots
                                all_lab=get_data["lab"].split(",")
                                    
                                for j in all_lab:
                                    for i in range(frm,to+1):
                                        mycursor.execute("SELECT `{0}` FROM `{1}` WHERE Day='{2}';".format(str(i),j,get_data['day']))
                                        lab=mycursor.fetchall()
                                        print(lab)

                                        print(len(lab[0][0]),lab[0][0])

                                        if(not (lab[0][0]=='--')):
                                            return jsonify({'class_occupied':"lab:{0} is already occupied in this slot range".format(j)})


                                for i in fac_split:
                                    #select for "--" in fac list

                                    for j in range(frm,to+1):

                                        mycursor.execute("SELECT `{0}` FROM `{1}` WHERE Day='{2}';".format(str(j),i,get_data['day']))

                                        cls=mycursor.fetchall()

                                        print(cls)

                                        print(len(cls[0][0]),cls[0][0])

                                        if(not cls[0][0]=='--'):  

                                            return jsonify({'fac_occupied':"faculty:{0} is already occupied in this slot range".format(i)})
                                        #endfor inner                             
                                #endfor

                                #update class
                                for i in range(frm,to+1):
                                    mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE Day='{2}';".format(str(i),get_data['class'],get_data['day'],get_data['subject']))
                                    mydb.commit()

                                    data.allocated=data.allocated+1
                                    db.session.commit()  

                                    if(data.dayperiod=="@"):
                                        data.dayperiod=get_data['day']+"--"+str(i)
                                        db.session.commit()
                                    else:
                                        data.dayperiod= data.dayperiod +","+get_data['day']+"--"+str(i)
                                        db.session.commit() 

                                #update list of lab
                                for i in all_lab:
                                    for j in range(frm,to+1):
                                        mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE Day='{2}';".format(str(j),i,get_data['day'],get_data['class']+'/'+get_data['subject']))
                                        mydb.commit()



                                #update faculty  list

                                for j in fac_split:

                                    faculty=Teacheradd.query.filter_by(name=j).first()
                                    
                                    for i in range(frm,to+1):
                                        mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE `Day`='{2}';".format(str(i),j,get_data['day'],get_data['class']+'/'+get_data['subject']))
                                        mydb.commit()     

                                        faculty.work_load=faculty.work_load+1
                                        db.session.commit()

                                #endfor 
                                
                                data.faculty=get_data['faculty']
                                db.session.commit()

                            else:
                                return jsonify({'day_error':"from_day should less than to_day"})

                        else:
                            print("min error")
                            return jsonify({'min_error':" faculty count exceeded"})
                    else:
                        return jsonify({'duplicate_faclty':"faculty should not be repeated"})
                else:#ALREADY SOMEONE IS THERE
                    
                    if(flag==0):

                        data_fac_split=data.faculty.split(",")

                        if(set(data_fac_split)==set(fac_split)):

                            if(frm<to+1):
                                #checking availabilty of class slots
                                for i in range(frm,to+1):
                                    mycursor.execute("SELECT `{0}` FROM `{1}` WHERE Day='{2}';".format(str(i),get_data['class'],get_data['day']))

                                    cls=mycursor.fetchall()
                                    print(cls)

                                    print(len(cls[0][0]),cls[0][0])

                                    if(not (cls[0][0]=='--')):  

                                        return jsonify({'class_occupied':"class {0} already occupied in this slot range".format(get_data['class'])})
                                #endfor

                                #checking availability in lab slots
                                all_lab=get_data["lab"].split(",")
                                    
                                for j in all_lab:
                                    for i in range(frm,to+1):
                                        mycursor.execute("SELECT `{0}` FROM `{1}` WHERE Day='{2}';".format(str(i),j,get_data['day']))
                                        lab=mycursor.fetchall()
                                        print(lab)

                                        print(len(lab[0][0]),lab[0][0])

                                        if(not (lab[0][0]=='--')):
                                            return jsonify({'class_occupied':"lab:{0} is already occupied in this slot range".format(j)})                         
                                #endfor

                                for i in fac_split:
                                    #select for "--" in fac list

                                    for j in range(frm,to+1):

                                        mycursor.execute("SELECT `{0}` FROM `{1}` WHERE Day='{2}';".format(str(j),i,get_data['day']))

                                        cls=mycursor.fetchall()

                                        print(cls)

                                        print(len(cls[0][0]),cls[0][0])

                                        if(not cls[0][0]=='--'):  

                                            return jsonify({'fac_occupied':"faculty is already occupied in this slot range"})
                                    #endfor inner                             
                                #endfor

                                #update class
                                for i in range(frm,to+1):

                                    mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE Day='{2}';".format(str(i),get_data['class'],get_data['day'],get_data['subject']))
                                    mydb.commit()


                                    data.allocated=data.allocated+1
                                    db.session.commit()  

                                    if(data.dayperiod=="@"):
                                        data.dayperiod=get_data['day']+"--"+str(i)
                                        db.session.commit()
                                    else:
                                        data.dayperiod= data.dayperiod +","+get_data['day']+"--"+str(i)
                                        db.session.commit()

                                for j in all_lab:

                                    for i in range(frm,to+1):

                                        mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE Day='{2}';".format(str(i),j,get_data['day'],get_data['class']+'/'+get_data['subject']))
                                        mydb.commit()

                                #update faculty  list

                                for j in fac_split:
                                    faculty=Teacheradd.query.filter_by(name=j).first()

                                    for i in range(frm,to+1):
                                        mycursor.execute("UPDATE `{1}` SET `{0}`='{3}' WHERE `Day`='{2}';".format(str(i),j,get_data['day'],get_data['class']+'/'+get_data['subject']))
                                        mydb.commit()     

                                        faculty.work_load=faculty.work_load+1
                                        db.session.commit() 

                                #endfor 
                            else:
                                return jsonify({'day_error':"from_day should less than to_day"})
                                
                        else:
                            return jsonify({'duplicate_faclty':"make sure of list of faculty assigned previously"})
                    else:
                        return jsonify({'duplicate_faclty':"faculty should not be repeated"})
            else:
                return jsonify({'max':"maximum allotment of periods satisfied"})


        #update split by appending '/' and scale
        scale=(int(get_data['to'])+1)-int(get_data['from'])
        if(data.allocated-scale==0):
            #append "/"

            data.split=data.split+"/"+str(scale)
            db.session.commit()
        else:
            data.split=data.split+","+str(scale)
            db.session.commit()



        mycursor.execute("SELECT * FROM `{0}`;".format(get_data['class']))
        row=mycursor.fetchall()
        all_data=classconfig_table.query.all()

        return jsonify( {'data': render_template('classtable.html',all_data=row,periods=get_period.periods),'describe':render_template('classtable_describe.html',classname=get_data['class'],all_data=all_data)})
    else:
        return redirect(url_for('dashboard'))


@app.route('/classwisetimetable')
def classwisetimetable():


    if 'loggedin' in session:

        if(session['type']=='faculty_type'):#

            print('dept_type is not dept_type')

            return redirect(url_for('faculty_table'))


        #welcome

        if(session['type']=='principal_type'):

            print("welcome to principal")
            all_class=classconfig_table.query.with_entities(classconfig_table.classname).distinct()

        else:

            print('welcome dept admin ')
        
            mycursor = mydb.cursor(buffered=True)

            dept=department.query.filter_by(mail=session['mail']).first()
            
            #all_class=classconfig_table.query.with_entities(classconfig_table.classname).distinct()

            all_class=[]
            class_from_course=Courseadd.query.filter_by(dept=dept.name).all()

            for cls in class_from_course:

                reg=cls.name+'%'
                print(reg)
        
                mycursor.execute("SHOW TABLES WHERE Tables_in_{0} LIKE '{1}'; ".format("jntuk1",reg))

                table=mycursor.fetchall()
                for i in table:

                    print("classname",i[0])

                    if(i[0] not in all_class):

                        all_class.append(i[0])

                reg=''
            
        all_classconfig=classconfig_table.query.all()

        all_dept=Teacheradd.query.with_entities(Teacheradd.branch).distinct()

        all_teacher=Teacheradd.query.all()
        list_total_faculty=[]
        for i in all_teacher:
                list_faculty=[]
                list_faculty.append(i.name)
                list_faculty.append(i.branch)

                list_total_faculty.append(list_faculty)


        t=[]
        for i in all_classconfig:
            l=[]
            l.append(i.id)
            l.append(i.subject)
            l.append(i.type)
            l.append(i.faculty)
            l.append(i.classname)
            l.append(i.lab)
            l.append(i.total_periods)
            l.append(i.f_count)
            l.append(i.allocated)
            l.append(i.dayperiod)
            t.append(l)

            #get day count and period count

        get_day_period=day_period.query.get(1)


        days=get_day_period.day
        periods=get_day_period.periods
        day_list=['MON','TUE','WED','THU','FRI','SAT','SUN']

        day_list=day_list[:days]

        print(days,periods)  


        return render_template("class_wise_timetable.html",all_class=all_class,all_classconfig=t,all_dept=all_dept,all_teacher=list_total_faculty,days=days,periods=periods,day_list=day_list,access=session['type'])
    else:
        return redirect(url_for('dashboard')) 


class classconfig_table(db.Model):

    __tablename__ = "classconfig_table"
    #Subject	Type	Teacher	Class	Lab Name	Total Periods 	Faculty_count
    id = db.Column(db.Integer, primary_key = True)
    subject= db.Column(db.String(100))
    type= db.Column(db.String(40))
    faculty=db.Column(db.String(200))
    classname=db.Column(db.String(200))
    lab=db.Column(db.String(200))
    total_periods=db.Column(db.Integer)
    f_count=db.Column(db.Integer)
    allocated=db.Column(db.Integer)
    dayperiod=db.Column(db.String(200))
    split=db.Column(db.String(100))



    def __init__(self,subject,type,faculty,classname,lab,total_periods,f_count,allocated,dayperiod,split):

        self.subject=subject
        self.type=type
        self.faculty=faculty
        self.classname,self.lab=classname,lab
        self.total_periods=total_periods
        self.f_count,self.allocated=f_count,allocated
        self.dayperiod=dayperiod
        self.split=split
        
db.create_all()


#This is the class configuration route where we are going to
@app.route('/classconfig')
def classconfig():
    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))

        if(session['type']=='principal_type'):
            all_course=Courseadd.query.all()   
        else:     

            dept=department.query.filter_by(mail=session['mail']).first()
            
            #all_class=classconfig_table.query.with_entities(classconfig_table.classname).distinct()

            all_course=Courseadd.query.filter_by(dept=dept.name).all()


        all_teacher=Teacheradd.query.all()

        all_subject=Subject.query.all()

        list_total_faculty=[]
        
        for i in all_teacher:
                list_faculty=[]
                list_faculty.append(i.name)
                list_faculty.append(i.branch)

                list_total_faculty.append(list_faculty)


        """for i in all_subject:

            print(type(i.name))"""

        """for i in all_subject:
                        print(i.name)"""

        all_lab=Labs.query.all()

        #print(type(all_lab))



        l1,l2=[],[]
        for i in all_course:
                l1.append(i.name)
                l2.append(i.semisters)
                
        pythontojson=dict(zip(l1,l2))


        """for i in all_subject:
                
                        print(i.name)"""

            #all_subject_list.append[list(i)]
            #print(all_subject_list,list(i))


        l2=[]
        for i in all_subject:
            
            l1=[]
            l1.append(i.id)
            l1.append(i.name)
            l1.append(i.course)
            l1.append(i.sem)
            l1.append(i.subjecttype)

            l2.append(l1)      
            
        #all_subject=map(json.dumps,l2)   
        return render_template("classconfig.html",all_course=all_course,
        all_dept=Teacheradd.query.with_entities(Teacheradd.branch).distinct(),
        all_teacher=list_total_faculty,all_subject=l2,
        json_object=pythontojson,all_lab=all_lab,access=session['type'])
    else:
	    return redirect(url_for('dashboard'))

@app.route('/send', methods = ['GET', 'POST'])
def send():
    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))

        if request.method=='POST':
            #print(request.get_json())
            jsondata = request.get_json()

            flag=show_all_tables((jsondata['table_name']).lower())

            get_day_period=day_period.query.get(1)
            
            if(flag is False):
                get_day_period=day_period.query.get(1)

                if(get_day_period):
                    print("creating class table:{0}".format(jsondata['table_name']))

                    create_table(jsondata['table_name'],day=get_day_period.day,periods=get_day_period.periods)

                    #mycursor = mydb.cursor(buffered=True)
                else:
                    flash("please set No of Days/Periods")

                    return redirect(url_for('edit_days_periods'))


            #print(jsondata['type'][:9])

            #get_allocated=session.query(classconfig_table).filter(classconfig_table.classname==jsondata['table_name'])

            get_allocated=classconfig_table.query.filter_by(classname=jsondata['table_name']).all()

            #get_type=classconfig_table.query.filter_by(classname=jsondata['table_name'],).all()

            mycursor.execute("SELECT DISTINCT type FROM classconfig_table WHERE type  LIKE 'ELECTIVE/%' AND `classname` LIKE '{0}' ".format(jsondata['table_name']))

            type_val=mycursor.fetchall()

            type_list=[]
            for j in type_val:
                type_list.append(j[0])

                print("type list",j[0])


            summ=0
            new_type_list=[]
            for j in get_allocated:

                if(j.type in type_list):
                    
                    if(j.type not in new_type_list):

                        new_type_list.append(j.type)

                        summ+=int(j.total_periods)
                else:

                    summ+=int(j.total_periods)


            print("summ",summ+int(jsondata['period']))

            if(summ+int(jsondata['period'])>int(get_day_period.day)*int(get_day_period.periods)):

                print("entered if")
                return jsonify({'distinct':"Maximum periods of allotment to this class is {0} ,total periods configured till now is {1}".format(get_day_period.day*get_day_period.periods,summ)}) 

            all_split=jsondata['split_p'].split(',')

            for sp in all_split:

                if(int(sp)>max(get_day_period.morning_periods , get_day_period.day-get_day_period.morning_periods)):

                    return jsonify({'distinct':"Maximum period split should less than equal to morning periods or afternoon periods "})
      

            if(len((jsondata['split_p']).split(','))<=get_day_period.day):
                
                #session.query(classconfig_table).filter(classconfig_table.classname==jsondata['table_name'],classconfig_table.subject==jsondata['subject']).count()

                val=classconfig_table.query.filter_by(classname=jsondata['table_name'],subject=jsondata['subject']).all()


                print("distinct",len(val)==0)

                if(len(val)==0):
                    

                    if(jsondata['type']== "THEORY" or jsondata['type'][:9]=="ELECTIVE/"):

                            faculty_split_role=(jsondata['faculty']).split(',')

                            for i in faculty_split_role:

                                get_fac=Teacheradd.query.filter_by(name=i).first()

                                designation=get_fac.role

                                print(designation)

                                #create dictionary matching old role with new role

                                designation_dic={'Professor':'proffesor','Associate Professor':'Assoc_prof','Assistant Professor':'Asst_prof','Assistant Professor(C)':'Asst_prof_c'}

                                mycursor.execute("SELECT {1} FROM `{0}` WHERE id=1".format("day_periods",designation_dic[designation]))

                                val=mycursor.fetchall()

                                print("designation",val[0][0])

                                print(get_fac.work_load+int(jsondata['period']))

                                if( (get_fac.work_load+int(jsondata['period'])> val[0][0] ) and (designation=="Professor" or designation=="Associate Professor" or designation=="Assistant Professor")):

                                    return jsonify({'distinct':"{0} is exceeding his/her maximum work_limit periods".format(jsondata['faculty'])})                
                    
                            add=classconfig_table(jsondata['subject'],jsondata['type'],
                            jsondata['faculty'],jsondata['table_name'],"NA",jsondata['period'],1,0,"@",jsondata['split_p'])

                            db.session.add(add)
                            db.session.commit()
                            
                            #print(all_tables)

                            #print( jsonify( {'data': render_template('response_table_classconf.html', myList=total_list)}))

                            all_data=classconfig_table.query.all()

                            return jsonify( {'data': render_template('response_table_classconf.html',all_data=all_data,table=jsondata['table_name'] )})

                        

                    elif(jsondata['type']=="LAB"):

                        add=classconfig_table(jsondata['subject'],jsondata['type'],
                        "NA",jsondata['table_name'],jsondata['labname'],jsondata['period'],jsondata['faculty_count'],0,"@",jsondata['split_p'])

                        db.session.add(add)
                        db.session.commit()

                        all_data=classconfig_table.query.all()
                        return jsonify( {'data': render_template('response_table_classconf.html',all_data=all_data,table=jsondata['table_name'] )})

                    else:

                        add=classconfig_table(jsondata['subject'],jsondata['type'],
                        "NA",jsondata['table_name'],"NA",jsondata['period'],jsondata['faculty_count'],0,"@",jsondata['split_p'])

                        db.session.add(add)
                        db.session.commit()
                        all_data=classconfig_table.query.all()
                        
                        return jsonify( {'data': render_template('response_table_classconf.html',all_data=all_data,table=jsondata['table_name'] )})

                else:
                    return jsonify({'distinct':"subjects should be unique under single class"})
                    #return jsonify( {'data': render_template('response_table_classconf.html',all_data=all_data,table=jsondata['table_name'] )},{'distinct':"subjects should be unique under single class"})
            else:
                return jsonify({'distinct':"Split period exceeding the out of Days"})
    else:
        return redirect(url_for('dashboard'))
@app.route('/delete_conf', methods = ['GET', 'POST'])
def delete_conf():
    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))


        ids = request.json['ids']
        table_class = request.json['table']

        check_all=request.json['check_all']

        print(ids)
        flag=1
        

        if(ids):
            if (',' in ids):
                ids = ids.split(',')

                for i in ids:

                    my_data = classconfig_table.query.get(i)

                    if(my_data.dayperiod =='@'):

                        db.session.delete(my_data)

                        db.session.commit()
                    else:

                        flag=0

            else:


                my_data = classconfig_table.query.get(ids)

                if(my_data.dayperiod =='@'):

                    db.session.delete(my_data)

                    db.session.commit()
                else:

                    flag=0

            if(check_all and flag):

                deleteF(table_class)
            all_data=classconfig_table.query.all()

            print('return')
            return jsonify( {'data': render_template('response_table_classconf.html',all_data=all_data,table = table_class )})

        else:
            
            return jsonify( {'data': render_template('response_table_classconf.html',all_data=all_data,table = table_class )})
    else:
        return redirect(url_for('dashboard'))
            
@app.route('/gettable', methods = ['GET', 'POST'])
def gettable():
    if 'loggedin' in session:
        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))


        if request.method=='POST':


            all_data=classconfig_table.query.all()

            print(request.get_json())

            jsondata=request.get_json()

            #return( jsonify( {'data':jsondata['table'] }))

            return jsonify( {'data': render_template('response_table_classconf.html',all_data=all_data,table=jsondata['table'])})
    else:
        return redirect(url_for('dashboard'))
@app.route('/edit_days_periods')
def edit_days_periods():
    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))


        get_day_period=day_period.query.get(1)

        if(get_day_period):

            return render_template("edit_days_periods.html",day=get_day_period.day,periods=get_day_period.periods,morning_periods=get_day_period.morning_periods,prof=get_day_period.proffesor,assoc=get_day_period.Assoc_prof,assis=get_day_period.Asst_prof,assis_c=get_day_period.Asst_prof_c,access=session['type'])
    else:
        return redirect(url_for('dashboard'))

@app.route('/reset_days_periods', methods = ['POST','GET'])
def reset_days_periods():
    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))


        get_data=request.get_json()

        get_day_period=day_period.query.get(1)

        if(get_day_period):

            print("something")
            print(get_data)
            old_day=get_day_period.day
            old_period=get_day_period.periods

            print("old day",old_day,"old_period",old_period)

            if(int(get_data['morning_periods']) > int(get_data['days']) ):

                return jsonify({'message':"FIRST HALF PERIODS IS EXCEEDING THE TOTAL PERIODS "})


            
            get_day_period.day=int(get_data['days'])
            db.session.commit() 
            get_day_period.periods=int(get_data['periods'])
            db.session.commit()

            #update workload also  and each workload should less than day*periods

            max_periods=get_day_period.day*get_day_period.periods


            get_day_period.proffesor=int(get_data['proff'])
            db.session.commit()
            get_day_period.Assoc_prof=int(get_data['assoc_proff'])
            db.session.commit()
            get_day_period.Asst_prof=int(get_data['assistent_proff'])
            db.session.commit()
            get_day_period.Asst_prof_c=int(get_data['assistent_proff_c'])
            db.session.commit()

            get_day_period.morning_periods=int(get_data['morning_periods'])
            db.session.commit()        

            new_day=get_day_period.day
            new_period=get_day_period.periods

            print("new day",new_day,"new_period",new_period)

            all_class=classconfig_table.query.with_entities(classconfig_table.classname).distinct()


            #excel_function

            flag=excel_report()


            for i in all_class:
                #alter/update
                deleteF(i.classname)

            # delete all classconfigdata
            delete_conf=classconfig_table.query.all()

            for i in delete_conf:
                get_id=classconfig_table.query.get(i.id)
                db.session.delete(get_id)
                db.session.commit()

            delete_conf=classconfig_table.query.all()

            if(delete_conf):
                print("something")
            else:
                print("nothing")

            # making workload=0
            
            all_faculty=Teacheradd.query.all()

            for i in all_faculty:
                update=Teacheradd.query.get(i.id)
                update.work_load=0
                db.session.commit()

                deleteF(i.name)

                if(not show_all_tables((i.name).lower())):
                    print(i.name,"passed")
                    create_table(i.name,int(get_data['periods']),int(get_data['days']))
                else:
                    print(i.name,"is not present")

            mycursor = mydb.cursor(buffered=True)
            all_lab=Labs.query.all()

            for i in all_lab:
                deleteF(i.name)

                if(show_all_tables):
                    create_table(i.name,int(get_data['periods']),int(get_data['days']))
            mycursor = mydb.cursor(buffered=True)
            print("worked")



            #change size of existing faculty/lab


        else:
            print("nothing")

            #default_day_period=day_period(day=get_data['days'],periods=get_data['periods'])
            #db.session.add(default_day_period)
            #db.session.commit()    

            #print(get_data)

        return jsonify({'message':"reset successful"})
    else:
        return redirect(url_for('dashboard'))
#Creating edit_day_periods table for our jntuk database
class day_period(db.Model):
    __tablename__ = "day_periods"
    id = db.Column(db.Integer, primary_key = True)
    day = db.Column(db.Integer)
    periods=db.Column(db.Integer)

    proffesor=db.Column(db.Integer)
    Assoc_prof=db.Column(db.Integer)
    Asst_prof=db.Column(db.Integer)
    Asst_prof_c=db.Column(db.Integer)
    morning_periods=db.Column(db.Integer)

    def __init__(self,day,periods,proff,associate_prof,assistent_prof,assistent_prof_c,morning_periods):

        self.day = day
        self.periods=periods

        self.proffesor=proff

        self.Assoc_prof=associate_prof

        self.Asst_prof=assistent_prof

        self.Asst_prof_c=assistent_prof_c

        self.morning_periods=morning_periods

db.create_all()



#Creating Teacheradd table for our jntuk database
class subjecttype(db.Model):
    __tablename__ = "subjecttype"

    id = db.Column(db.Integer, primary_key = True)
    name = db.Column(db.String(100))

    def __init__(self, name):
        self.name = name

db.create_all()

@app.route('/subjecttype')
def subjecttypef():

    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))


        all_data = subjecttype.query.all()

        return render_template("subjecttype.html",dept=all_data,access=session['type'])
    else:
        return redirect(url_for('dashboard'))

#this route is for inserting btech subjects data to mysql database via html forms

@app.route('/subjecttype_add', methods = ['POST'])
def subjecttype_add():

    if 'loggedin' in session:  

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))

    
        if request.method == 'POST':
    
            name = request.form['name']

            all_data = subjecttype.query.all()

            for i in all_data:

                if(name==i.name):

                    flash("Already present")
                    return redirect(url_for('subjecttypef'))


            my_data = subjecttype(name)
            db.session.add(my_data)
            db.session.commit()

        flash("subjecttype Added Successfully")

        return redirect(url_for('subjecttypef'))
    else:
        return redirect(url_for('dashboard'))

@app.route('/subjecttype_edit/<id>', methods = ['POST'])
def subjecttype_edit(id):
    if 'loggedin' in session: 
        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))
        
        if request.method == 'POST':
            my_data = subjecttype.query.get(id)
            
            all_data = subjecttype.query.all()

            for i in all_data:

                if(request.form['name']==i.name):

                    flash("Already present")
                    return redirect(url_for('subjecttypef'))

            my_data.name = request.form['name']
            db.session.commit()

        flash("subjecttype Added Successfully")

        return redirect(url_for('subjecttypef'))
    else:
        return redirect(url_for('dashboard'))

#This route is for deleting our lab
@app.route('/subjecttype_delete/<id>/', methods = ['GET', 'POST'])
def subjecttype_delete(id):
    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))

        my_data = subjecttype.query.get(id)
        db.session.delete(my_data)
        db.session.commit()
        flash("subjecttype Deleted Successfully")
    
        return redirect(url_for('subjecttypef'))
    else:
        return redirect(url_for('dashboard'))
    
#Creating Teacheradd table for our jntuk database
class department(db.Model):
    __tablename__ = "department"
    id = db.Column(db.Integer, primary_key = True)
    name = db.Column(db.String(100))
    mail= db.Column(db.String(100))

    def __init__(self, name,mail):
        self.name = name
        self.mail=mail

db.create_all()

@app.route('/department')
def dept():
    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))


        all_data = department.query.all()

        return render_template("department.html",dept=all_data,access=session['type'] )
    else:
        return redirect(url_for('dashboard'))
#this route is for inserting btech subjects data to mysql database via html forms




#This route is for deleting our lab
@app.route('/dept_delete/<id>/', methods = ['GET', 'POST'])
def dept_delete(id):
    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))

        my_data = department.query.get(id)

        username=my_data.mail

        db.session.delete(my_data)
        db.session.commit()

        my_login = login.query.filter_by(username=username).first()

        id=login.query.get(my_login.id)


        db.session.delete(id)

        db.session.commit()

        flash("Department Deleted Successfully")
    
        return redirect(url_for('dept'))
    else:
        return redirect(url_for('dashboard'))
            
#Creating Teacheradd table for our jntuk database
class Teacheradd(db.Model):

    __tablename__ = "Facultyadd"
    id = db.Column(db.Integer, primary_key = True)
    name = db.Column(db.String(100))
    branch = db.Column(db.String(20))
    role=db.Column(db.String(50))

    mail=db.Column(db.String(100))
    work_load= db.Column(db.Integer)
    exclude=db.Column(db.Integer)


    def __init__(self, name,branch,role,mail):
        self.name = name
        self.branch=branch
        self.role=role
        self.mail=mail
        self.work_load=0
        self.exclude=0

        
db.create_all()



#query on all our teacher data
@app.route('/teacheradd')
def teacheradd():
    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))

        #all_data=session.query(Teacheradd).order_by(desc(Teacheradd.id)).all() # desc
        all_data = Teacheradd.query.all()

        all_data=Teacheradd.query.order_by(Teacheradd.work_load).all()

        #all_data.reverse()

        all_dept=department.query.all()

        return render_template("teacheradd.html", employees = all_data,count=len(all_data),all_dept=all_dept,access=session['type'])
    else:
        return redirect(url_for('dashboard'))
#this route is for inserting btech subjects data to mysql database via html forms

@app.route('/dept_add', methods = ['POST','GET'])
def dept_add():
    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))

        if request.method == 'POST':
    
            name = request.form['name']

            mail= request.form['mail']

            all_data = department.query.all()

            for i in all_data:

                if(name==i.name):

                    flash(" Department Already exists")
                    return redirect(url_for('dept'))
                if(mail==i.mail):
                    flash("mail is already exists")
                    return redirect(url_for('dept'))

            password='dept@jntuk'

            print("password",generate_password_hash(password))



            login_data=login(mail,generate_password_hash(password),"dept_type",0)
            db.session.add(login_data)
            db.session.commit()


            my_data = department(name,mail)
            db.session.add(my_data)
            db.session.commit()

        flash("Department Added Successfully")

        return redirect(url_for('dept'))
    else:
        return redirect(url_for('dashboard'))
@app.route('/dept_edit/<id>/<mail_old>', methods = ['POST'])
def dept_edit(id,mail_old):
    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))

        if request.method == 'POST':
            my_data = department.query.get(id)
            
            all_data = department.query.all()
            flag=1
            for i in all_data:

                if(request.form['name']==i.name):

                    flash(" Dept Already exists")

                    flag=0

            if(flag==1):
                my_data.name = request.form['name']
                db.session.commit()

            flag=0
            for i in all_data:

                if(request.form['mail']==i.mail):
                    flash("mails Already exists")
                    flag=1
                    

            if(flag==0):
                print(mail_old)

                my_login = login.query.filter_by(username=mail_old).first()

                print(my_login.username)

                id=login.query.get(my_login.id)

                id.username=request.form['mail']

                db.session.commit()  

                print(id.username)
                flash("mail updated")

                my_data.mail=request.form['mail']
                db.session.commit()

        #flash("Department Updated Successfully")

        return redirect(url_for('dept'))
    else:
        return redirect(url_for('dashboard'))



#this route is for inserting teacher data to mysql database via html forms
@app.route('/insert', methods = ['POST'])
def insert():
    if 'loggedin' in session:
        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))

        if request.method == 'POST':

            mycursor = mydb.cursor(buffered=True)
    
            name = request.form['name'].upper()#challa
            #name=name.replace(" ","_")

            if('na'==name):
                flash('faculty name should not be NA ')

                return redirect(url_for('teacheradd'))

            count=Teacheradd.query.filter_by(name=name).count()

            mail = request.form['mail']
            #name=name.replace(" ","_")
            count2=Teacheradd.query.filter_by(mail=mail).count()        

            if(count==0 and count2==0):
                branch=request.form['branch']

                role=request.form['role']
            
                my_data = Teacheradd(name,branch,role,mail)
                db.session.add(my_data)
                db.session.commit()#insert


                password='faculty@jntuk'
                print("pasword",generate_password_hash(password))

                login_data=login(mail,generate_password_hash(password),"faculty_type",0)
                db.session.add(login_data)
                db.session.commit()


            else:
                flash("Faculty/mail is already exists")


                return redirect(url_for('teacheradd'))
                

            flag=show_all_tables(name.lower())

            if(flag is False):

                get_day_period=day_period.query.get(1)

                if(get_day_period):
                    print("something")

                    create_table(name,day=get_day_period.day,periods=get_day_period.periods)
                    mycursor = mydb.cursor(buffered=True)
                    flash("Faculty Table created Successfully")
                else:
                    flash("please set No of Days/Periods")

            return redirect(url_for('teacheradd'))
    else:
        return redirect(url_for('dashboard'))   
    
#this is our update route where we are going to update our Faculty
@app.route('/updateteacher/<name1>/<mail_old>', methods = ['GET', 'POST'])
def updateteacher(name1,mail_old):
    if 'loggedin' in session:
    
        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))

        if request.method == 'POST':
            my_data = Teacheradd.query.get(request.form.get('id'))

            name=request.form['name'].upper() 

            mail = request.form['mail']

            count2=Teacheradd.query.filter_by(mail=mail).count() 

            if(count2==0):#update mail
            
                print(mail_old)

                my_login = login.query.filter_by(username=mail_old).first()

                print(my_login.username)

                id=login.query.get(my_login.id)

                id.username=mail

                db.session.commit()  

                print(id.username)          

                my_data.mail=mail
                db.session.commit()#insert


            my_data.branch=request.form['branch']
            my_data.role=request.form['role']
            db.session.commit()#insert
            #name.replace(" ","_")
            flash("Updations Effected")

            flag=show_all_tables(name.lower())

            if(flag is False):

                all_in_conf=classconfig_table.query.all()

                for fac in all_in_conf:

                    fac_split=fac.faculty.split(',')

                    if(name1 in fac_split):

                        ids=classconfig_table.query.get(fac.id)

                        fac_split.remove(name1)
                        fac_split.append(name)

                        ids.faculty=(',').join(fac_split)

                        db.session.commit()

                #name.replace(" ","_")
                my_data.name=name
                db.session.commit()#insert

                updateF(name1,name)#table name


            
                flash("Employee Updated successfully")
            else:
                flash("Employee is present under this name")
            return redirect(url_for('teacheradd'))
    else:
        return redirect(url_for('dashboard'))
 
 
#This route is for deleting our faculty
@app.route('/deleteteacher/<id>/', methods = ['GET', 'POST'])
def deleteteacher(id):
    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))

        my_data = Teacheradd.query.get(id)

        flag=show_all_tables((my_data.name).lower())

        if(flag is True):

            if(my_data.work_load!=0):
                flash("Faculty is still configured somewhere in else Set faculty Free")
                return redirect(url_for('teacheradd'))

            deleteF(my_data.name)

            username=my_data.mail

            db.session.delete(my_data)

            db.session.commit()

            my_login = login.query.filter_by(username=username).first()

            id=login.query.get(my_login.id)

            db.session.delete(id)

            db.session.commit()

            flash("Employee Deleted Successfully")
    
        return redirect(url_for('teacheradd'))
    else:
        return redirect(url_for('dashboard'))

#creating configure teacher/subjects table
class Teacherconfigtable(db.Model):

    __tablename__ = "teacherconfigtable"
    id = db.Column(db.Integer, primary_key = True)

    name = db.Column(db.String(100))

    subject=db.Column(db.String(100))
 
    def __init__(self, name,subject):
 
        self.name = name

        self.subject=subject


db.create_all()




#This is the class configuration route where we are going to
@app.route('/teacherconfig',methods=["POST", "GET"])
def teacherconfig():
    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))


        """mycursor = mydb.cursor()
                
                    mycursor.execute("SELECT name from teacheradd")
                
                    myresult = mycursor.fetchall()
                
                    res=[list(i) for i in myresult]"""

        res2=Teacheradd.query.all()

        """for i in range(len(res)):
                        res2.append(res[i][0])"""

        sub=Subject.query.with_entities(Subject.name).all()

        conf=Teacherconfigtable.query.all()

        for i in sub:
            print(i)

        return render_template("teacherconfigworking.html", teacherlis=res2,sub=sub,conf=conf,access=session['type'])
    else:
        return redirect(url_for('dashboard'))


#configuring teachers and subjects
@app.route('/teacherconfigadd',methods=["POST", "GET"])    
def teacherconfigadd():
    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))

        if request.method == 'POST':
     
            name = request.form['teachers'].upper()
            #name=name.replace(" ","_")
            subject=request.form['subjects']
            my_data = Teacherconfigtable(name,subject)
            db.session.add(my_data)
            db.session.commit()
    
            flash("subject is configured to faculty Successfully")
    
            return redirect(url_for('teacherconfig'))
    else:
        return redirect(url_for('dashboard'))
    


#Creating Classadd table for our jntuk database
class Courseadd(db.Model):
    __tablename__ = "courseadd"
    id = db.Column(db.Integer, primary_key = True)

    name = db.Column(db.String(250))

    semisters=db.Column(db.Integer)

    dept= db.Column(db.String(250))


 
    def __init__(self, coursename,semisters,dept):
 
        self.name = coursename

        self.semisters=semisters

        self.dept=dept


db.create_all()


@app.route('/courseadd1')
def courseadd():
    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))

        all_data = Courseadd.query.all()

        all_dept=department.query.all()
    
        return render_template("courseadd.html", courses= all_data,all_dept=all_dept,access=session['type'])
    else:
        return redirect(url_for('dashboard'))

#all routes of dropdown CLassadd

@app.route('/insertcourse', methods = ['POST','GET'])
def insertcourse():
    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))

        if request.method == 'POST':
            coursename = request.form['course'].upper()
            coursename=coursename.replace(" ","_")
            semisters=request.form['semisters']

            dept=request.form['dept']
            count=Courseadd.query.filter_by(name=coursename).count()

            if('/' in coursename):
                flash("course name is not allowed with '/' symbol ")
                return redirect(url_for('courseadd'))

            if((len(coursename.split('-')))!=3):

                flash("coursename format should be course - regulation - branch,should contain exactly 2 '-' ")
                return redirect(url_for('courseadd'))

            all_course=Courseadd.query.all()

            for all in all_course:

                if(coursename in all.name or all.name in coursename):
                    flash("coursename should not be any substring of existing of courses ",'danger')
                    return redirect(url_for('courseadd'))

            if(count==0 ):

                my_data = Courseadd(coursename,semisters,dept)
                db.session.add(my_data)
                db.session.commit()
                flash("course inserted successfully ")

            else:
                flash("course is already present ")

    
            return redirect(url_for('courseadd'))
    else:
        return redirect(url_for('dashboard'))
    
#this is our update route where we are going to update our Faculty
@app.route('/updatecourse/<name1>', methods = ['GET', 'POST'])
def updatecourse(name1):
    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))

        if request.method == 'POST':
            my_data = Courseadd.query.get(request.form.get('id'))

            coursename= request.form['course'].upper()
            coursename=coursename.replace(" ","_")
            
            my_data.semisters =request.form['semisters']

            db.session.commit()
            

            my_data.dept=request.form['dept']
            db.session.commit()



            flash("Updations Affected")

            all_course=Courseadd.query.all()

            for all in all_course:

                if(coursename in all.course or all.course in coursename):
                    flash("coursename should not be any substring of existing of courses ")
                    return redirect(url_for('courseadd'))

            if(Courseadd.query.filter_by(name=coursename.replace(" ","_")).count()==0):
                coursename=coursename.replace(" ","_")


                mycursor = mydb.cursor(buffered=True)

                print("------------------>",coursename)
                reg=name1+'%'
                print(reg)
            
                mycursor.execute("SHOW TABLES WHERE Tables_in_{0} LIKE '{1}'; ".format("jntuk1",reg))

                table=mycursor.fetchall()
                #alter all table names

                for i in table:

                    print("old table:",i[0])

                    if(name1 in i[0]):

                        print("new table name:",i[0].replace(name1,coursename))
                    mycursor.execute("ALTER TABLE `{0}` RENAME TO `{1}`".format(i[0],i[0].replace(name1,coursename)))

                    mydb.commit()
                
                #alter classconfig_table
                   
                mycursor.execute(" SELECT id FROM classconfig_table WHERE classname LIKE '{0}';".format(reg))
                classconfig=mycursor.fetchall()
                for i in classconfig:
                    print(" Alter TABLE:",i[0])
                    my_data1 = classconfig_table.query.get(i[0])

                    if(name1 in my_data1.classname):
                        #repalce it
                        my_data1.classname=my_data1.classname.replace(name1,coursename)
                        db.session.commit()

                #alter Subjects       
                mycursor.execute(" SELECT id FROM subject WHERE course='{0}';".format(name1))
                alter_sub=mycursor.fetchall()
                for i in alter_sub:
                    print(" in subjects TABLE:",i[0])
                    my_data1 = Subject.query.get(i[0])

                    my_data1.course=coursename

                    db.session.commit()

                my_data.name=coursename

                db.session.commit()
                flash("course Updated Successfully")

                return redirect(url_for('courseadd'))

            else:
                flash("course already present")
    
            return redirect(url_for('courseadd'))
    else:
        return redirect(url_for('dashboard'))
 
 

#This route is for deleting our employee
@app.route('/deletecourse/<id>/<name>', methods = ['GET', 'POST'])
def deletecourse(id,name):
    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))

        print(id)
        my_data = Courseadd.query.get(id)
        #print(my_data.id,my_data.name)

        deletec(name)
        db.session.delete(my_data)
        db.session.commit()

    
        flash("course Deleted Successfully")
    
        return redirect(url_for('courseadd'))
    else:
        return redirect(url_for('dashboard'))    

def deletec(table_name):

    if 'loggedin' in session: 
        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))

        mycursor = mydb.cursor(buffered=True)

        print("------------------>",table_name)
        reg=table_name+'%'
        print(reg)
    
        mycursor.execute("SHOW TABLES WHERE Tables_in_{0} LIKE '{1}'; ".format("jntuk1",reg))

        table=mycursor.fetchall()

        for i in table:

            print(i[0])

            flag=class_reset(i[0])
            if(flag):

                mycursor.execute("DROP TABLE `{0}` ".format(i[0]))
                mydb.commit()

        #if some x course is deleted than classes/semisters of that course will get deleted    
        mycursor.execute(" SELECT id FROM classconfig_table WHERE classname LIKE '{0}';".format(reg))
        for i in mycursor:
            print("TABLE:",i[0])
            my_data = classconfig_table.query.get(i[0])
            db.session.delete(my_data)
            db.session.commit()


        #if some x course is deleted than subjects of that course will get deleted       
        mycursor.execute(" SELECT id FROM subject WHERE course='{0}';".format(table_name))
        for i in mycursor:
            print(" in subjects TABLE:",i[0])
            my_data = Subject.query.get(i[0])
            db.session.delete(my_data)
            db.session.commit()
    else:
        return redirect(url_for('dashboard'))

#Creating btechsubject table for our jntuk database
class Subject(db.Model):
    #__tablename__ = "subjectadd"

    id = db.Column(db.Integer, primary_key = True)

    name=  db.Column(db.String(100))

    course= db.Column(db.String(60))

    dept= db.Column(db.String(60))

    sem= db.Column(db.String(10))

    subjecttype= db.Column(db.String(20))

 
    def __init__(self,name,sem,subjecttype,course,dept):
        
        #name.replace(" ","_")
        self.name = name
        self.course=course
        self.subjecttype=subjecttype
        self.sem=sem
        self.dept=dept


db.create_all()


#route on all our btech subjects list
@app.route('/subject')
def subject():
    if 'loggedin' in session:
        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))

        #all_subjects = Subject.query.all()

        if(session['type']=='principal_type'): 
            all_subjects = Subject.query.all()
            all_course=Courseadd.query.all()
        else:

            dept=department.query.filter_by(mail=session['mail']).first()
            
            #all_class=classconfig_table.query.with_entities(classconfig_table.classname).distinct()

            all=Courseadd.query.filter_by(dept=dept.name).all()

            all_subjects=[]

            for co in all:

                print(co.name)

                all_subjects+=Subject.query.filter_by(course=co.name).all()



            all_course=Courseadd.query.filter_by(dept=dept.name).all()



        #print(all_course)

        #mycursor = mydb.cursor()

        #mycursor.execute("SELECT name,semisters from courseadd")

        #all_course = mycursor.fetchall()





        l1,l2=[],[]
        for i in all_course:
                l1.append(i.name)
                l2.append(i.semisters)
                
        pythontojson=dict(zip(l1,l2))

        """for i,j in all_courses.items():
                print(i,j)"""
                
        json_object = json.dumps(pythontojson)
        #print(json_object)


        sem=8
        #r=db.engine.execute('SELECT name,semisters from courseadd')

        all_dept=department.query.all()

        all_type=subjecttype.query.all()
    
        return render_template("subject.html",subjects = all_subjects,all_course=all_course,sem=sem,json_object=pythontojson,all_dept=all_dept,all_type=all_type,access=session['type'])
    else:
        return redirect(url_for('dashboard')) 


#this route is for inserting btech subjects data to mysql database via html forms
@app.route('/subjectadd', methods = ['POST'])
def subjectadd():
    if 'loggedin' in session:
        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))
    
        if request.method == 'POST':
    
            name = request.form['name'].upper()


            name_list=name.split(",")

            name_list=[i.strip() for i in name_list]

            for i in name_list:
                #i.replace(" ","_")
                name=i

                course=request.form['course']
                sem=request.form['sem']
                dept=request.form['dept']
                subjecttype=request.form['subjecttype']

                #session.query(Subject).filter(Subject.name==name,Subject.course==course).count()

                #print(session.query(Subject).filter(Subject.name==name,Subject.course==course).count())

                

                if(Subject.query.filter_by(name=name,course=course).count()==0):

                    my_data = Subject(name,sem,subjecttype,course,dept)
                    db.session.add(my_data)
                    db.session.commit()
                    

                else:
                    flash("{0} subject under this course already present ".format(name))
                    #return redirect(url_for('subject'))
    
            flash("subject Inserted Successfully")
    
            return redirect(url_for('subject'))
    else:
        return redirect(url_for('dashboard'))

#this is our update route where we are going to update our subject
@app.route('/subjectedit', methods = ['GET', 'POST'])
def subjectedit():
    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))
        if request.method == 'POST':
            my_data = Subject.query.get(request.form.get('id'))
    
            name= request.form['name'].upper()
            #name.replace(" ","_")
            #name=name


            print("changing fac_dept",request.form['dept'])
            my_data.dept=request.form['dept']
            db.session.commit()
    
            my_data.sem= request.form['sem']
            db.session.commit()
    
            my_data.subjecttype= request.form['subjecttype']

            db.session.commit()


            #result = session.query(Subject).filter(Subject.course==request.form['course']).count()

            #print(request.form['course'],Subject.query.filter_by(course=request.form['course']).count(),result)
            
            #sub=Subject.query.filter_by(course=request.form['course']).all()
            #session.query(Subject).filter(Subject.course==request.form['course'],Subject.name==name).count()

            print(Subject.query.filter_by(course=request.form['course'],name=name).count())

            if(Subject.query.filter_by(course=request.form['course'],name=name).count()==0):
                my_data.name=name
                db.session.commit()
                my_data.course=request.form['course']
                db.session.commit()
            else:
                flash("subject under this course already present")

                return redirect(url_for('subject'))


            flash("subject Updated Successfully")
    
            return redirect(url_for('subject'))
    else:
        return redirect(url_for('dashboard'))

#This route is for deleting our subject
@app.route('/deletesubject', methods = ['GET', 'POST'])
def deletesubject():
    if 'loggedin' in session:
        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))
        ids = request.json['ids']
        
        print(ids)
        if(ids):
            if (',' in ids):
                ids = ids.split(',')

                for i in ids:

                    my_data = Subject.query.get(i)

                    db.session.delete(my_data)

                    db.session.commit()

                return jsonify("true")


            else:
                my_data = Subject.query.get(ids)

                db.session.delete(my_data)

                db.session.commit()

                #return redirect(url_for('subject'))

                return jsonify("false")
        
                #flash("subject Deleted Successfully")
    else:
        return redirect(url_for('dashboard'))
    



 
#Creating labadd table for our jntuk database
class Labs(db.Model):
    __tablename__ = "labs"
    id = db.Column(db.Integer, primary_key = True)
    name = db.Column(db.String(100),unique=True)
    branch = db.Column(db.String(10))

    def __init__(self, name,branch):
        self.name = name
        self.branch=branch

db.create_all()


#query on all our lab data
@app.route('/labpage')
def labpage():
    if 'loggedin' in session:

        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))
        all_data = Labs.query.all()

        for i in all_data:

            print(i)

        all_dept=department.query.all()


        return render_template("labadd.html", labs = all_data,count=len(all_data),all_dept=all_dept,access=session['type'])
        
    else:
        return redirect(url_for('dashboard'))

#this route is for inserting teacher data to mysql database via html forms
@app.route('/labadd', methods = ['POST'])
def labadd():
    if 'loggedin' in session:
        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))
    
        if request.method == 'POST':
    
            name = request.form['name'].upper()
            
            #name=name.replace(" ","_")
            branch=request.form['branch']

            count=Labs.query.filter_by(name=name).count()

            my_data = Labs(name,branch)
            if(count==0):
                db.session.add(my_data)
                db.session.commit()
                flag=show_all_tables(name.lower())

                if(flag is False):

                    get_day_period=day_period.query.get(1)

                    if(get_day_period):
                        print("something")

                        create_table(name,day=get_day_period.day,periods=get_day_period.periods)

                        mycursor = mydb.cursor(buffered=True)

                        flash("Lab Inserted Successfully")
                    else:
                        flash("please set No of Days/Periods")

                else:
                    flash("Lab is already present ")
            else:
                flash("Lab  Already  present")
    
            return redirect(url_for('labpage'))
    else:
        return redirect(url_for('dashboard'))
 
#this is our update route where we are going to update our lab
@app.route('/updatelab/<name1>', methods = ['GET', 'POST'])
def updatelab(name1):
    if 'loggedin' in session:
        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))
        if request.method == 'POST':
            my_data = Labs.query.get(request.form.get('id'))
    
            name = request.form['name'].upper()

            my_data.branch=request.form['branch']


            #name=name.replace(" ","_")
            
            flag=show_all_tables(name.lower())

            if(flag is False):
                
                my_data.name=name.replace(" ","_")
                db.session.commit()
                updateF(name1,my_data.name)

                all_lab=classconfig_table.query.all()

                for lab in all_lab:
                    lab_split=lab.lab.split(',')

                    if(name1 in lab.lab.split(',')):

                        ids=classconfig_table.query.get(lab.id)

                        lab_split.remove(name1)
                        lab_split.append(name)

                        ids.lab=(',').join(lab_split)

                        db.session.commit()
                        
                flash("Lab updated Successfully")
            else:
                flash("Lab already present")


            return redirect(url_for('labpage'))
    else:
        return redirect(url_for('dashboard'))
#This route is for deleting our lab
@app.route('/deletelab/<id>/', methods = ['GET', 'POST'])
def deletelab(id):
    if 'loggedin' in session:
        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))

        my_data = Labs.query.get(id)
        db.session.delete(my_data)
        db.session.commit()
        deleteF(my_data.name)
        flash("Lab Deleted Successfully")
    
        return redirect(url_for('labpage'))
    else:
        return redirect(url_for('dashboard'))

#This is the lab configuration route where we are going to
@app.route('/labconfig')
def labconfig():
    if 'loggedin' in session:
        if(session['type']=='faculty_type'):

            return redirect(url_for('faculty_table'))
        all_course=Courseadd.query.all()

        all_teacher=Teacheradd.query.all()

        all_subject=Subject.query.all()

        l1,l2=[],[]
        for i in all_course:
                l1.append(i.name)
                l2.append(i.semisters)
                
        pythontojson=dict(zip(l1,l2))

        return render_template("labconfig.html",all_course=all_course,all_teacher=all_teacher,all_subject=all_subject,json_object=pythontojson,access=session['type'])
    else:
        return redirect(url_for('dashboard'))
if __name__ == "__main__":
    app.run(debug=True)